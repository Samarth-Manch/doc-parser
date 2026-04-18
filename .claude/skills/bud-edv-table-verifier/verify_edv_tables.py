#!/usr/bin/env python3
"""Audit EDV reference tables in a BUD .docx.

Verifies that every uppercase-underscore table name cited in "EDV" field logic
is actually present as an embedded Excel object under section "4.6" of the doc.
Emits a self-contained HTML report.

Exit codes:
  0 — every referenced name is present under 4.6
  1 — at least one name is missing or found only outside 4.6
  2 — docx could not be opened / parsed
"""

from __future__ import annotations

import argparse
import html
import os
import re
import sys
import tempfile
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from xml.etree import ElementTree as ET

NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "o": "urn:schemas-microsoft-com:office:office",
    "v": "urn:schemas-microsoft-com:vml",
    "rels": "http://schemas.openxmlformats.org/package/2006/relationships",
}

W_P = f"{{{NS['w']}}}p"
W_TBL = f"{{{NS['w']}}}tbl"
W_TC = f"{{{NS['w']}}}tc"
W_T = f"{{{NS['w']}}}t"
W_TAB = f"{{{NS['w']}}}tab"
W_BR = f"{{{NS['w']}}}br"
W_PSTYLE = f"{{{NS['w']}}}pStyle"
W_VAL = f"{{{NS['w']}}}val"
R_ID = f"{{{NS['r']}}}id"

NAME_RE = re.compile(r"\b([A-Z][A-Z0-9]*(?:_[A-Z0-9]+)+)\b")
EDV_CONTEXT_RE = re.compile(
    r"\b(EDV|Staging|staging|reference\s+table|lookup|validate\s+against|attribute\s+\d+\s+of)\b",
    re.IGNORECASE,
)
HEADING_46_RE = re.compile(r"^\s*4\.6(?:[.\s)]|$)")
HEADING_NUM_RE = re.compile(r"^\s*(\d+)\.(\d+)?")


# ---------- data types ----------

@dataclass
class Reference:
    name: str
    citations: list[tuple[str, str]] = field(default_factory=list)  # (section, snippet)


@dataclass
class Embedding:
    rid: str
    target: str              # word/embeddings/...
    filename: str
    label: str               # preceding paragraph text
    sheets: list[str]
    headers_preview: list[str]
    in_section_46: bool


@dataclass
class AuditResult:
    docx_path: str
    section_46_found: bool
    references: list[Reference]
    embeddings: list[Embedding]
    name_status: dict[str, dict]  # name -> {status, embedding_idx}


# ---------- helpers ----------

def _para_text(p: ET.Element) -> str:
    parts: list[str] = []
    for el in p.iter():
        tag = el.tag
        if tag == W_T and el.text:
            parts.append(el.text)
        elif tag == W_TAB:
            parts.append("\t")
        elif tag == W_BR:
            parts.append("\n")
    return "".join(parts).strip()


def _para_style(p: ET.Element) -> str | None:
    ppr = p.find(f"{{{NS['w']}}}pPr")
    if ppr is None:
        return None
    style = ppr.find(W_PSTYLE)
    if style is None:
        return None
    return style.get(W_VAL)


def _is_heading(p: ET.Element) -> bool:
    style = _para_style(p) or ""
    if style.lower().startswith("heading"):
        return True
    txt = _para_text(p)
    return bool(re.match(r"^\s*\d+(\.\d+)*\.?\s+\S", txt)) and len(txt) < 200


def _collect_rids(el: ET.Element) -> list[str]:
    rids: list[str] = []
    for sub in el.iter():
        rid = sub.get(R_ID)
        if rid:
            rids.append(rid)
    return rids


def _table_cell_texts(tbl: ET.Element) -> list[str]:
    cells: list[str] = []
    for tc in tbl.iter(W_TC):
        bits = []
        for p in tc.iter(W_P):
            t = _para_text(p)
            if t:
                bits.append(t)
        cells.append(" ".join(bits))
    return cells


def _parse_rels(rels_xml: bytes) -> dict[str, str]:
    root = ET.fromstring(rels_xml)
    out: dict[str, str] = {}
    for rel in root.findall(f"{{{NS['rels']}}}Relationship"):
        out[rel.get("Id")] = rel.get("Target")
    return out


def _normalize_rel_target(target: str) -> str:
    # rels targets are relative to word/, e.g. "embeddings/foo.xlsx"
    if target.startswith("/"):
        return target[1:]
    if target.startswith("word/"):
        return target
    return "word/" + target


def _read_xlsx_meta(data: bytes) -> tuple[list[str], list[str]]:
    """Return (sheet_names, first_nonempty_headers_of_first_sheet)."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return [], []
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(data)
            path = tmp.name
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            sheets = list(wb.sheetnames)
            headers: list[str] = []
            if sheets:
                ws = wb[sheets[0]]
                for row in ws.iter_rows(max_rows=1, values_only=True):
                    headers = [str(c) if c is not None else "" for c in row]
                    break
            wb.close()
            return sheets, headers
        finally:
            try:
                os.unlink(path)
            except OSError:
                pass
    except Exception:
        return [], []


# ---------- main audit ----------

def audit(docx_path: str) -> AuditResult:
    with zipfile.ZipFile(docx_path, "r") as z:
        doc_xml = z.read("word/document.xml")
        try:
            rels_xml = z.read("word/_rels/document.xml.rels")
        except KeyError:
            rels_xml = b"<Relationships xmlns='http://schemas.openxmlformats.org/package/2006/relationships'/>"
        namelist = z.namelist()
        embed_data = {
            n: z.read(n)
            for n in namelist
            if n.startswith("word/embeddings/")
        }

    rid_to_target = _parse_rels(rels_xml)
    root = ET.fromstring(doc_xml)
    body = root.find(f"{{{NS['w']}}}body")
    if body is None:
        raise ValueError("docx has no body")

    # Walk body children linearly. Track 4.6 window and current heading.
    in_46 = False
    section_46_found = False
    current_section = "(preamble)"
    last_para_text = ""   # used as label for subsequent embedding

    references: dict[str, Reference] = {}
    embeddings: list[Embedding] = []

    def record_embeddings(rids: list[str], label: str, in_46: bool) -> None:
        for rid in rids:
            target = rid_to_target.get(rid)
            if not target:
                continue
            norm = _normalize_rel_target(target)
            if not norm.startswith("word/embeddings/"):
                continue
            filename = norm.rsplit("/", 1)[-1]
            ext = filename.lower().rsplit(".", 1)[-1]
            sheets: list[str] = []
            headers: list[str] = []
            if ext in ("xlsx", "xlsm"):
                data = embed_data.get(norm)
                if data is not None:
                    sheets, headers = _read_xlsx_meta(data)
            embeddings.append(Embedding(
                rid=rid,
                target=norm,
                filename=filename,
                label=label,
                sheets=sheets,
                headers_preview=headers,
                in_section_46=in_46,
            ))

    def scan_for_references(text: str, section: str) -> None:
        if not text or not EDV_CONTEXT_RE.search(text):
            return
        for m in NAME_RE.finditer(text):
            name = m.group(1)
            if name in {"EDV", "NA", "N_A"}:
                continue
            ref = references.setdefault(name, Reference(name=name))
            snippet = text.strip()
            if len(snippet) > 240:
                snippet = snippet[:237] + "…"
            ref.citations.append((section, snippet))

    for child in body:
        tag = child.tag
        if tag == W_P:
            txt = _para_text(child)
            if _is_heading(child) and txt:
                current_section = txt
                if HEADING_46_RE.match(txt):
                    in_46 = True
                    section_46_found = True
                elif in_46:
                    m = HEADING_NUM_RE.match(txt)
                    if m:
                        major = int(m.group(1))
                        minor = int(m.group(2)) if m.group(2) else None
                        # leave 4.6 on encountering 4.7+ or 5./6./...
                        if major > 4 or (major == 4 and minor is not None and minor > 6):
                            in_46 = False
            rids = _collect_rids(child)
            if rids:
                label = last_para_text or current_section
                record_embeddings(rids, label, in_46)
            scan_for_references(txt, current_section)
            if txt:
                last_para_text = txt
        elif tag == W_TBL:
            # Scan cell text for references
            cells = _table_cell_texts(child)
            for cell in cells:
                scan_for_references(cell, current_section)
            # Also any embeddings inside the table
            rids = _collect_rids(child)
            if rids:
                label = last_para_text or current_section
                record_embeddings(rids, label, in_46)

    # Classify each referenced name.
    name_status: dict[str, dict] = {}
    for name in references:
        match_in = None
        match_out = None
        up = name.upper()
        for idx, emb in enumerate(embeddings):
            hay = " ".join([emb.label, emb.filename, " ".join(emb.sheets)]).upper()
            if up in hay:
                if emb.in_section_46:
                    match_in = idx
                    break
                if match_out is None:
                    match_out = idx
        if match_in is not None:
            name_status[name] = {"status": "present", "embedding_idx": match_in}
        elif match_out is not None:
            name_status[name] = {"status": "outside_4_6", "embedding_idx": match_out}
        else:
            name_status[name] = {"status": "missing", "embedding_idx": None}

    return AuditResult(
        docx_path=docx_path,
        section_46_found=section_46_found,
        references=sorted(references.values(), key=lambda r: r.name),
        embeddings=embeddings,
        name_status=name_status,
    )


# ---------- HTML rendering ----------

STATUS_BADGE = {
    "present": ("#1b7a1b", "PRESENT"),
    "outside_4_6": ("#a06a00", "OUTSIDE 4.6"),
    "missing": ("#b00020", "MISSING"),
}


def render_html(res: AuditResult) -> str:
    total = len(res.references)
    present = sum(1 for n in res.name_status.values() if n["status"] == "present")
    outside = sum(1 for n in res.name_status.values() if n["status"] == "outside_4_6")
    missing = sum(1 for n in res.name_status.values() if n["status"] == "missing")

    emb_in = [e for e in res.embeddings if e.in_section_46]
    matched_emb_idx = {
        ns["embedding_idx"]
        for ns in res.name_status.values()
        if ns["embedding_idx"] is not None
    }
    orphans = [
        (i, e) for i, e in enumerate(res.embeddings)
        if e.in_section_46 and i not in matched_emb_idx
    ]

    def esc(s: str) -> str:
        return html.escape(s or "")

    def badge(status: str) -> str:
        color, label = STATUS_BADGE[status]
        return (
            f'<span style="background:{color};color:#fff;padding:2px 8px;'
            f'border-radius:10px;font-size:11px;font-weight:600;">{label}</span>'
        )

    rows_ref = []
    for ref in res.references:
        ns = res.name_status[ref.name]
        idx = ns["embedding_idx"]
        match_desc = "—"
        if idx is not None:
            e = res.embeddings[idx]
            match_desc = esc(f"{e.filename}  |  sheets: {', '.join(e.sheets) or 'n/a'}")
        cites = "<br>".join(
            f"<code>{esc(sec)}</code>: {esc(sn)}"
            for sec, sn in ref.citations[:4]
        )
        if len(ref.citations) > 4:
            cites += f"<br><em>+{len(ref.citations) - 4} more</em>"
        rows_ref.append(
            f"<tr><td><code>{esc(ref.name)}</code></td>"
            f"<td>{badge(ns['status'])}</td>"
            f"<td>{match_desc}</td>"
            f"<td>{cites}</td></tr>"
        )

    rows_inv = []
    for e in emb_in:
        rows_inv.append(
            f"<tr><td><code>{esc(e.filename)}</code></td>"
            f"<td>{esc(e.label)}</td>"
            f"<td>{esc(', '.join(e.sheets) or '—')}</td>"
            f"<td>{esc(', '.join(h for h in e.headers_preview if h) or '—')}</td></tr>"
        )

    rows_orph = []
    for _, e in orphans:
        rows_orph.append(
            f"<tr><td><code>{esc(e.filename)}</code></td>"
            f"<td>{esc(e.label)}</td>"
            f"<td>{esc(', '.join(e.sheets) or '—')}</td></tr>"
        )

    missing_names = [n for n, s in res.name_status.items() if s["status"] == "missing"]
    outside_names = [n for n, s in res.name_status.items() if s["status"] == "outside_4_6"]

    banner = ""
    if not res.section_46_found:
        banner = (
            '<div style="background:#fff4e5;border:1px solid #e0a84d;'
            'padding:12px;border-radius:6px;margin:16px 0;">'
            '<strong>⚠ No heading matching "4.6" was detected in this document.</strong> '
            'Every referenced name will therefore appear as missing or outside-4.6.'
            '</div>'
        )

    ok_banner = ""
    if total == 0:
        ok_banner = (
            '<div style="background:#e8f5e9;border:1px solid #4caf50;'
            'padding:12px;border-radius:6px;margin:16px 0;">'
            'No EDV table references were found in the document logic. Nothing to verify.'
            '</div>'
        )

    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>EDV Table Audit — {esc(os.path.basename(res.docx_path))}</title>
<style>
  body {{ font-family: -apple-system, Segoe UI, Roboto, sans-serif; margin: 24px; color:#222; max-width: 1200px; }}
  h1 {{ font-size: 22px; margin: 0 0 4px; }}
  .meta {{ color:#666; font-size: 13px; margin-bottom: 20px; }}
  .cards {{ display:flex; gap: 12px; margin: 16px 0 24px; }}
  .card {{ flex:1; border:1px solid #e1e4e8; border-radius:8px; padding:14px; background:#fafbfc; }}
  .card .n {{ font-size: 28px; font-weight: 700; }}
  .card .l {{ font-size: 12px; text-transform: uppercase; color:#586069; letter-spacing: 0.5px; }}
  .card.ok .n {{ color:#1b7a1b; }}
  .card.warn .n {{ color:#a06a00; }}
  .card.err .n {{ color:#b00020; }}
  h2 {{ font-size: 16px; border-bottom: 1px solid #eaecef; padding-bottom: 6px; margin-top: 32px; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
  th, td {{ text-align:left; padding: 8px 10px; border-bottom: 1px solid #eaecef; vertical-align: top; }}
  th {{ background:#f6f8fa; font-weight:600; }}
  code {{ font-family: SFMono-Regular, Consolas, monospace; background:#f6f8fa; padding:1px 4px; border-radius:3px; font-size: 12px; }}
  .missing-list {{ background:#ffebee; border:1px solid #f5c6cb; padding:10px 14px; border-radius:6px; }}
  .missing-list code {{ background:#fff; }}
</style></head><body>
<h1>EDV Reference Table Audit</h1>
<div class="meta">
  <strong>Document:</strong> {esc(os.path.basename(res.docx_path))}<br>
  <strong>Path:</strong> <code>{esc(res.docx_path)}</code><br>
  <strong>Generated:</strong> {esc(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}<br>
  <strong>Section 4.6 located:</strong> {'yes' if res.section_46_found else 'no'}
</div>

{banner}{ok_banner}

<div class="cards">
  <div class="card"><div class="n">{total}</div><div class="l">Referenced names</div></div>
  <div class="card ok"><div class="n">{present}</div><div class="l">Present in 4.6</div></div>
  <div class="card warn"><div class="n">{outside}</div><div class="l">Outside 4.6</div></div>
  <div class="card err"><div class="n">{missing}</div><div class="l">Missing</div></div>
</div>

{("<h2>Referenced but Missing</h2><div class='missing-list'>" + ", ".join(f"<code>{esc(n)}</code>" for n in missing_names) + "</div>") if missing_names else ""}

{("<h2>Found Outside 4.6</h2><div class='missing-list' style='background:#fff8e1;border-color:#e0a84d;'>" + ", ".join(f"<code>{esc(n)}</code>" for n in outside_names) + "</div>") if outside_names else ""}

<h2>Referenced names</h2>
<table>
  <thead><tr><th>Name</th><th>Status</th><th>Matched embedding</th><th>Cited in (section · snippet)</th></tr></thead>
  <tbody>{''.join(rows_ref) or '<tr><td colspan=4><em>No EDV references found.</em></td></tr>'}</tbody>
</table>

<h2>Excel embeddings under 4.6 ({len(emb_in)})</h2>
<table>
  <thead><tr><th>File</th><th>Label</th><th>Sheets</th><th>First-row headers</th></tr></thead>
  <tbody>{''.join(rows_inv) or '<tr><td colspan=4><em>No embedded Excel objects under section 4.6.</em></td></tr>'}</tbody>
</table>

<h2>Orphan embeddings in 4.6 ({len(orphans)})</h2>
<p style="color:#666;font-size:12px;">Excel objects inside 4.6 that did not match any name referenced in logic.</p>
<table>
  <thead><tr><th>File</th><th>Label</th><th>Sheets</th></tr></thead>
  <tbody>{''.join(rows_orph) or '<tr><td colspan=3><em>None.</em></td></tr>'}</tbody>
</table>

</body></html>
"""


# ---------- CLI ----------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("docx", help="Path to the BUD .docx file")
    parser.add_argument(
        "--output",
        default="output/edv_table_audit",
        help="Directory for the HTML report (default: output/edv_table_audit)",
    )
    args = parser.parse_args(argv)

    if not os.path.isfile(args.docx):
        print(f"error: file not found: {args.docx}", file=sys.stderr)
        return 2

    try:
        result = audit(args.docx)
    except zipfile.BadZipFile:
        print("error: file is not a valid .docx (zip) archive", file=sys.stderr)
        return 2
    except ET.ParseError as e:
        print(f"error: malformed document.xml: {e}", file=sys.stderr)
        return 2

    os.makedirs(args.output, exist_ok=True)
    stem = os.path.splitext(os.path.basename(args.docx))[0]
    out_path = os.path.join(args.output, f"{stem}.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(render_html(result))

    total = len(result.references)
    present = sum(1 for s in result.name_status.values() if s["status"] == "present")
    outside = sum(1 for s in result.name_status.values() if s["status"] == "outside_4_6")
    missing = sum(1 for s in result.name_status.values() if s["status"] == "missing")
    print(f"Report: {out_path}")
    print(f"Referenced: {total}  Present: {present}  Outside 4.6: {outside}  Missing: {missing}")

    return 0 if (missing == 0 and outside == 0 and total > 0) else (0 if total == 0 else 1)


if __name__ == "__main__":
    sys.exit(main())
