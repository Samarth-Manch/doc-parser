#!/usr/bin/env python3
"""
Script to create External Data Value (EDV) metadata on the Manch platform
from an Excel or CSV file.

Usage:
    python3 create_edv.py --auth-token <TOKEN> --excel data.xlsx [--sheet <SHEET>]
    python3 create_edv.py --auth-token <TOKEN> --excel data.csv
    python3 create_edv.py --auth-token <TOKEN> --excel data.xlsx --name MY_EDV --dry-run

The script:
1. Reads the Excel/CSV file, extracts column headers as EDV attributes
2. Analyzes the data to determine candidate key (uniqueness criteria)
3. Creates the EDV metadata via the Manch API

API: POST /app/v2/company/{companyId}/external-data-metadata
"""

import argparse
import csv
import json
import re
from itertools import combinations

import openpyxl
import requests


# ──────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────
BASE_URL = "https://qa.manchtech.com"
DEFAULT_COMPANY_ID = 248


# ──────────────────────────────────────────────
# File Reading (Excel + CSV)
# ──────────────────────────────────────────────

def _read_csv(file_path: str) -> tuple:
    """Read a CSV file and return (headers, data_rows, source_label)."""
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        # Sniff delimiter (comma, semicolon, tab, pipe)
        sample = f.read(8192)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel  # fallback to comma-separated

        reader = csv.reader(f, dialect)
        all_rows = list(reader)

    if not all_rows:
        raise ValueError(f"CSV file '{file_path}' is empty")

    headers = [h.strip() for h in all_rows[0]]
    data = [[c.strip() for c in row] for row in all_rows[1:]]

    # Pad short rows to match header length
    for i, row in enumerate(data):
        if len(row) < len(headers):
            data[i] = row + [""] * (len(headers) - len(row))
        elif len(row) > len(headers):
            data[i] = row[: len(headers)]

    return headers, data, file_path


def _read_excel(file_path: str, sheet_name: str = None) -> tuple:
    """Read an Excel file and return (headers, data_rows, source_label)."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
        sheet_name = wb.sheetnames[0]

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise ValueError(f"Excel sheet '{sheet_name}' is empty")

    headers = [str(h).strip() if h else "" for h in all_rows[0]]
    data = [[str(c).strip() if c else "" for c in row] for row in all_rows[1:]]

    return headers, data, sheet_name


def read_file(file_path: str, sheet_name: str = None) -> tuple:
    """Read an Excel or CSV file and return (headers, data_rows).

    Detects format by file extension (.csv vs .xlsx/.xls).

    Returns:
        (headers: list[str], rows: list[list[str]])
    """
    lower = file_path.lower()
    if lower.endswith(".csv"):
        headers, data, label = _read_csv(file_path)
        print(f"  Format: CSV")
    elif lower.endswith((".xlsx", ".xls")):
        headers, data, label = _read_excel(file_path, sheet_name)
        print(f"  Sheet: {label}")
    else:
        raise ValueError(
            f"Unsupported file format: {file_path}. Use .xlsx, .xls, or .csv"
        )

    # Remove trailing empty columns
    while headers and not headers[-1]:
        headers.pop()
        data = [row[: len(headers)] for row in data]

    # Remove fully empty rows
    data = [row for row in data if any(cell.strip() for cell in row)]

    print(f"  Columns: {len(headers)}, Data rows: {len(data)}")
    return headers, data


# ──────────────────────────────────────────────
# Candidate Key Detection
# ──────────────────────────────────────────────

# Column names that strongly suggest key/identifier columns
KEY_INDICATORS = re.compile(
    r"\b(id|code|key|number|no|num|identifier|ref|reference|vendor.?code|"
    r"employee.?id|emp.?id|item.?code|sku|ean|upc|gstin|pan|cin|"
    r"company.?code|plant.?code|material.?code)\b",
    re.IGNORECASE,
)

# Column names that suggest NON-key columns (descriptions, flags, etc.)
NON_KEY_INDICATORS = re.compile(
    r"\b(description|desc|name|message|status|flag|type|category|group|"
    r"remarks|comment|note|created|updated|modified|date|time|"
    r"phone|email|address|street|city|region|country|postal|pin|"
    r"fax|mobile|telephone)\b",
    re.IGNORECASE,
)


def _col_uniqueness_ratio(data: list, col_idx: int) -> tuple:
    """Return (unique_count, non_empty_count, ratio) for a column."""
    values = [row[col_idx] for row in data if row[col_idx].strip()]
    if not values:
        return 0, 0, 0.0
    unique = len(set(values))
    return unique, len(values), unique / len(values)


def _combo_is_unique(data: list, col_indices: tuple) -> bool:
    """Check if a combination of columns forms a unique key across all rows."""
    seen = set()
    for row in data:
        combo = tuple(row[i] for i in col_indices)
        if combo in seen:
            return False
        seen.add(combo)
    return True


def _col_key_score(header: str, uniqueness_ratio: float, col_idx: int) -> float:
    """Score a column's likelihood of being part of a candidate key.

    Higher score = more likely to be a key column.
    """
    score = 0.0

    # Uniqueness ratio is the strongest signal
    score += uniqueness_ratio * 50

    # Column name heuristics
    if KEY_INDICATORS.search(header):
        score += 30
    if NON_KEY_INDICATORS.search(header):
        score -= 20

    # Prefer earlier columns (slight bias)
    score -= col_idx * 0.5

    return score


def find_candidate_key(headers: list, data: list, max_key_size: int = 3) -> list:
    """Determine the candidate key (minimal set of columns that uniquely identify rows).

    Algorithm:
    1. Compute uniqueness ratio for each column
    2. Check single columns: if any column is fully unique AND looks like a key, use it
    3. Otherwise, try pairs sorted by key-likelihood score
    4. If no pair works, try triples
    5. Fallback: use all columns with highest uniqueness ratios

    Args:
        headers: Column header names
        data: Data rows
        max_key_size: Maximum number of columns in the candidate key

    Returns:
        List of 1-based column indices forming the candidate key
    """
    if not data:
        return [1]

    n_cols = len(headers)
    total_rows = len(data)

    # Step 1: Compute per-column stats
    col_stats = []
    for i in range(n_cols):
        unique, non_empty, ratio = _col_uniqueness_ratio(data, i)
        score = _col_key_score(headers[i], ratio, i)
        col_stats.append({
            "idx": i,
            "header": headers[i],
            "unique": unique,
            "non_empty": non_empty,
            "ratio": ratio,
            "score": score,
        })

    print("\n  Column analysis:")
    for cs in col_stats:
        indicator = ""
        if cs["ratio"] == 1.0 and cs["non_empty"] == total_rows:
            indicator = " ** UNIQUE"
        elif cs["ratio"] == 1.0 and cs["non_empty"] > 0:
            indicator = " * unique (sparse)"
        print(
            f"    [{cs['idx']+1}] {cs['header']}: "
            f"{cs['unique']}/{cs['non_empty']} unique "
            f"(ratio={cs['ratio']:.2f}, score={cs['score']:.1f}){indicator}"
        )

    # Step 2: Single-column candidate keys
    # A single column is a candidate key if it's fully unique across ALL rows
    single_candidates = [
        cs for cs in col_stats
        if cs["unique"] == total_rows and cs["non_empty"] == total_rows
    ]

    if single_candidates:
        # Among single unique columns, pick the one with highest key score
        best = max(single_candidates, key=lambda cs: cs["score"])
        result = [best["idx"] + 1]
        print(f"\n  Candidate key (single column): [{headers[best['idx']]}]")
        return result

    # Step 3: Try pairs of columns, prioritized by combined score
    # Only consider columns that have >1 unique value (exclude constants)
    viable = [cs for cs in col_stats if cs["unique"] > 1 and cs["non_empty"] > 0]
    viable.sort(key=lambda cs: cs["score"], reverse=True)
    # Limit search to top candidates for efficiency
    top_candidates = viable[:min(len(viable), 15)]

    for size in range(2, max_key_size + 1):
        for combo in combinations(top_candidates, size):
            indices = tuple(cs["idx"] for cs in combo)
            if _combo_is_unique(data, indices):
                result = sorted([idx + 1 for idx in indices])
                col_names = [headers[idx] for idx in sorted(indices)]
                print(f"\n  Candidate key ({size} columns): {col_names}")
                return result

    # Step 4: Fallback — use columns with highest uniqueness scores
    # Pick the top 2 columns by score that have good uniqueness
    fallback = sorted(viable, key=lambda cs: cs["score"], reverse=True)[:2]
    if fallback:
        result = sorted([cs["idx"] + 1 for cs in fallback])
        col_names = [headers[cs["idx"]] for cs in fallback]
        print(f"\n  Candidate key (fallback, best scores): {col_names}")
        return result

    # Ultimate fallback
    print("\n  Candidate key: using first column as fallback")
    return [1]


# ──────────────────────────────────────────────
# Attribute Name Sanitization
# ──────────────────────────────────────────────

def sanitize_attribute_name(name: str) -> str:
    """Convert a column header to a valid EDV attribute name.

    Rules: uppercase, replace spaces/special chars with underscores,
    remove leading/trailing underscores.
    """
    s = name.strip().upper()
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    s = s.strip("_")
    return s or "UNNAMED"


# ──────────────────────────────────────────────
# API Function
# ──────────────────────────────────────────────

def create_edv(
    base_url: str,
    auth_token: str,
    company_id: int,
    edv_name: str,
    edv_description: str,
    attributes: list,
    uniqueness_indices: list,
    dry_run: bool = False,
) -> dict:
    """Create an External Data Value metadata entry.

    Args:
        base_url: Base URL of the Manch platform
        auth_token: x-authorization token from login cookie
        company_id: Company ID from login session
        edv_name: Name/type of the EDV
        edv_description: Description text
        attributes: List of dicts with 'name' and 'mandatory' keys
        uniqueness_indices: 1-based indices of attributes for uniqueness criteria
        dry_run: If True, print payload but don't call API

    Returns:
        API response dict with status, message, id
    """
    attrs = {}
    for i, attr in enumerate(attributes, start=1):
        key = f"attribute{i}"
        mandatory_str = "true" if attr.get("mandatory") else "false"
        validation_msg = (
            f"Please provide value for {attr['name']}" if attr.get("mandatory") else None
        )
        attrs[key] = {
            "name": attr["name"],
            "displayOrder": i,
            "externalDataRules": [
                {
                    "ruleType": "VALIDATION",
                    "type": "MANDATORY",
                    "value": mandatory_str,
                    "validationMessage": validation_msg,
                }
            ],
            "edited": False,
        }

    criterias = [f"attribute{idx}value" for idx in uniqueness_indices]

    payload = {
        "externalDataType": edv_name,
        "description": edv_description,
        "genericData": "true",
        "criterias": criterias,
        "attributes": attrs,
    }

    headers = {
        "Content-Type": "application/json",
        "x-authorization": auth_token,
        "x-requested-with": "XMLHttpRequest",
    }

    url = f"{base_url}/app/v2/company/{company_id}/external-data-metadata"

    print(f"\n{'='*60}")
    print(f"EDV Name: {edv_name}")
    print(f"Description: {edv_description}")
    print(f"URL: POST {url}")
    print(f"Attributes ({len(attributes)}):")
    for i, a in enumerate(attributes, 1):
        is_key = i in uniqueness_indices
        print(f"  {i}. {a['name']}{' [UNIQUENESS]' if is_key else ''}"
              f"{' (mandatory)' if a.get('mandatory') else ''}")
    print(f"Uniqueness criteria: {criterias}")

    if dry_run:
        print(f"\n[DRY RUN] Payload:\n{json.dumps(payload, indent=2)}")
        return {"status": "DRY_RUN", "payload": payload}

    resp = requests.post(url, json=payload, headers=headers)
    resp.raise_for_status()
    result = resp.json()

    if result.get("status") == "SUCCESS":
        print(f"\nSUCCESS! EDV ID: {result.get('id')}")
    else:
        print(f"\nResponse: {result}")

    return result


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Create EDV on Manch platform from an Excel or CSV file. "
        "Reads column headers as attributes and auto-detects candidate key."
    )
    parser.add_argument(
        "--auth-token", required=True,
        help="x-authorization token (from login cookie's authToken field)",
    )
    parser.add_argument(
        "--excel", "--input", "-i", required=True, dest="input_file",
        help="Path to Excel (.xlsx) or CSV (.csv) file",
    )
    parser.add_argument(
        "--sheet", default=None,
        help="Sheet name for Excel files (default: first sheet; ignored for CSV)",
    )
    parser.add_argument(
        "--name", default=None,
        help="EDV name (default: derived from sheet name)",
    )
    parser.add_argument(
        "--description", default=None,
        help="EDV description (default: derived from sheet name)",
    )
    parser.add_argument(
        "--company-id", type=int, default=DEFAULT_COMPANY_ID,
        help=f"Company ID (default: {DEFAULT_COMPANY_ID})",
    )
    parser.add_argument(
        "--base-url", default=BASE_URL,
        help=f"Base URL (default: {BASE_URL})",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Print payload without calling the API",
    )

    args = parser.parse_args()

    print("=" * 60)
    print("Manch Platform - Create EDV from Excel/CSV")
    print("=" * 60)

    # Step 1: Read file
    print(f"\n[1] Reading: {args.input_file}")
    headers, data = read_file(args.input_file, args.sheet)

    # Step 2: Find candidate key
    print(f"\n[2] Analyzing candidate key...")
    uniqueness_indices = find_candidate_key(headers, data)

    # Step 3: Build attributes from headers
    attributes = []
    for i, header in enumerate(headers):
        attr_name = sanitize_attribute_name(header)
        is_key = (i + 1) in uniqueness_indices
        attributes.append({"name": attr_name, "mandatory": is_key})

    # Derive EDV name from sheet or filename if not specified
    if not args.name:
        edv_name = sanitize_attribute_name(
            args.sheet or headers[0] if headers else "UNNAMED"
        )
    else:
        edv_name = args.name

    description = args.description or f"EDV created from {args.input_file}"

    # Step 4: Create EDV
    print(f"\n[3] Creating EDV...")
    result = create_edv(
        base_url=args.base_url,
        auth_token=args.auth_token,
        company_id=args.company_id,
        edv_name=edv_name,
        edv_description=description,
        attributes=attributes,
        uniqueness_indices=uniqueness_indices,
        dry_run=args.dry_run,
    )

    print("\n" + "=" * 60)
    if not args.dry_run:
        print(f"Full response: {json.dumps(result, indent=2)}")
    print("Done!")


if __name__ == "__main__":
    main()
