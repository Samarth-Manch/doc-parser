"""
excel_writer.py
Shared Excel-writing helpers used by individual validators, plus the
registry-driven write_validation_report() entry point.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ── Status colours ─────────────────────────────────────────────────────────

SEVERITY_FILLS = {
    "FAIL":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "WARNING": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "PASS":    PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "INFO":    PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "N/A":     PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}

BOLD = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


# ── Helpers (used by validator write_sheet methods) ──────────────────────────

def write_header(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def apply_severity_fill(cell, value: str) -> None:
    fill = SEVERITY_FILLS.get(value)
    if fill:
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)


def write_pass_row(ws, row: int, cols: int, message: str) -> None:
    cell = ws.cell(row=row, column=1, value=message)
    cell.font = Font(bold=True)
    apply_severity_fill(cell, "PASS")


# ── Public API ───────────────────────────────────────────────────────────────

def write_validation_report(output_path: str, results: dict[str, list]) -> None:
    """Write all validation results to a single Excel workbook via the registry."""
    from validators.registry import ValidatorRegistry

    wb = Workbook()
    ValidatorRegistry.write_all(wb, results)
    wb.save(output_path)
