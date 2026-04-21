"""
Checks that ARRAY_HEADER (or ARRAY_HDR) and ARRAY_END in the Field Type
column of field sections are properly paired.
"""

from collections import defaultdict
from models import ArrayBracketCheckResult
from section_parser import FIELD_SECTIONS, FieldRow


OPEN_TOKENS = {"ARRAY_HEADER", "ARRAY_HDR"}
CLOSE_TOKEN = "ARRAY_END"


def _check_brackets_in_table(
    section: str, rows: list[FieldRow],
) -> list[ArrayBracketCheckResult]:
    """Validate ARRAY_HEADER/ARRAY_END pairing in a list of field rows."""
    results: list[ArrayBracketCheckResult] = []

    depth = 0
    open_row = None
    open_field = None

    for f in rows:
        if f.field_type_upper in OPEN_TOKENS:
            if depth >= 1:
                results.append(ArrayBracketCheckResult(
                    section=section, status="FAIL",
                    field_name=f.name, row=f.row_index,
                    message=f"Nested {f.field_type_upper} before closing previous array",
                    details=(
                        f"ARRAY_HEADER opened at row {open_row} "
                        f"(Field: '{open_field}') but no ARRAY_END before "
                        f"this {f.field_type_upper}. Nesting depth must not exceed 1."
                    ),
                    suggestion="Please close the previous array with ARRAY_END before opening a new one.",
                ))
            depth += 1
            open_row = f.row_index
            open_field = f.name

        elif f.field_type_upper == CLOSE_TOKEN:
            if depth == 0:
                results.append(ArrayBracketCheckResult(
                    section=section, status="FAIL",
                    field_name=f.name, row=f.row_index,
                    message="ARRAY_END without a preceding ARRAY_HEADER",
                    details="Found ARRAY_END but no ARRAY_HEADER / ARRAY_HDR was opened before it.",
                    suggestion="Please add the matching ARRAY_HDR before this ARRAY_END, or remove this ARRAY_END.",
                ))
            else:
                if open_field and f.name and open_field.strip().lower() != f.name.strip().lower():
                    results.append(ArrayBracketCheckResult(
                        section=section, status="FAIL",
                        field_name=f.name, row=f.row_index,
                        message="ARRAY_END field name does not match opening ARRAY_HEADER",
                        details=(
                            f"Opening ARRAY_HEADER at row {open_row} has field name "
                            f"'{open_field}' but this ARRAY_END has field name '{f.name}'."
                        ),
                        suggestion="Please make sure ARRAY_END has the same field name as its matching ARRAY_HDR.",
                    ))
                depth -= 1
                open_row = None
                open_field = None

    if depth > 0:
        results.append(ArrayBracketCheckResult(
            section=section, status="FAIL",
            field_name=open_field or "", row=open_row or 0,
            message="Unclosed ARRAY_HEADER \u2014 missing ARRAY_END",
            details=(
                f"ARRAY_HEADER opened at row {open_row} "
                f"(Field: '{open_field}') was never closed with ARRAY_END."
            ),
            suggestion="Please add the matching ARRAY_END to close this ARRAY_HDR.",
        ))

    # ── Name-based pairing check ──
    hdr_fields: dict[str, int] = {}
    end_fields: dict[str, int] = {}

    for f in rows:
        name_key = f.name_lower
        if f.field_type_upper in OPEN_TOKENS:
            hdr_fields[name_key] = f.row_index
        elif f.field_type_upper == CLOSE_TOKEN:
            end_fields[name_key] = f.row_index

    for name_key, r_idx in hdr_fields.items():
        if name_key not in end_fields:
            results.append(ArrayBracketCheckResult(
                section=section, status="FAIL",
                field_name=name_key, row=r_idx,
                message="ARRAY_HEADER has no ARRAY_END with the same field name",
                details=(
                    f"ARRAY_HEADER at row {r_idx} (Field: '{name_key}') "
                    f"has no corresponding ARRAY_END with the same field name."
                ),
                suggestion="Please add an ARRAY_END with the same field name to close this ARRAY_HDR.",
            ))

    return results


def validate_array_brackets(parsed) -> list[ArrayBracketCheckResult]:
    """Check ARRAY_HEADER/ARRAY_END pairing in field sections."""
    results: list[ArrayBracketCheckResult] = []

    for section_key in FIELD_SECTIONS:
        section = parsed.sections.get(section_key)
        if section is None or not section.heading_found:
            continue

        # Group field_rows by table_index for per-table checking
        tables: dict[int, list[FieldRow]] = defaultdict(list)
        for f in section.field_rows:
            tables[f.table_index].append(f)

        for t_idx in sorted(tables.keys()):
            results.extend(_check_brackets_in_table(section_key, tables[t_idx]))

    return results


# ── Registry integration ────────────────────────────────────────────────────

from openpyxl import Workbook
from .registry import BaseValidator, ValidatorRegistry, ValidationContext
from excel_writer import write_header, apply_severity_fill, auto_width, write_pass_row


@ValidatorRegistry.register
class ArrayBracketValidator(BaseValidator):
    name = "Validate Array Bracket Pairing"
    sheet_name = "Array Bracket Check"
    description = "Checks that ARRAY_HEADER and ARRAY_END markers in the Field Type column are properly paired."

    def validate(self, ctx: ValidationContext) -> list[ArrayBracketCheckResult]:
        return validate_array_brackets(ctx.parsed)

    def write_sheet(self, wb: Workbook, results: list[ArrayBracketCheckResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        headers = ["Section", "Field Name", "Row", "Message", "Details", "Status", "Suggestion"]
        write_header(ws, headers)

        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.section)
            ws.cell(row=i, column=2, value=r.field_name)
            ws.cell(row=i, column=3, value=r.row)
            ws.cell(row=i, column=4, value=r.message)
            ws.cell(row=i, column=5, value=r.details)
            ws.cell(row=i, column=6, value=r.status)
            apply_severity_fill(ws.cell(row=i, column=6), r.status)
            ws.cell(row=i, column=7, value=r.suggestion)

        if not results:
            write_pass_row(ws, 2, len(headers), "PASS - All ARRAY_HEADER/ARRAY_END pairs are properly matched.")

        auto_width(ws)
