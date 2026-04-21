"""
Checks that fields derived from other fields specify clear-on-change behavior.
"""

import re
from models import ClearFieldLogicResult


DEPENDENCY_PATTERNS = [
    re.compile(r'\b(?:derived|fetched?|look\s*up(?:ped)?)\s+from\b', re.IGNORECASE),
    re.compile(r'\bpopulate[sd]?\s+(?:based\s+on|from|depending\s+on)\b', re.IGNORECASE),
    re.compile(r'\b(?:dropdown|drop[-\s]?down)\s+values?\b[\s\S]*?\bbased\s+on\b', re.IGNORECASE),
    re.compile(r'\bvalues?\s+(?:are|from)\s+[\s\S]*?\b(?:column|table|edv)\b', re.IGNORECASE),
    re.compile(r'\bbased\s+on\b[\s\S]*?\bfield\b[\s\S]*?\bvalue\b', re.IGNORECASE),
    re.compile(r'\bbased\s+on\b[\s\S]*?\bvalue\b[\s\S]*?\bfield\b', re.IGNORECASE),
    re.compile(r'\bconcatenate[sd]?\s+with\b', re.IGNORECASE),
    re.compile(r'\bthrough\s+validation\s+of\b', re.IGNORECASE),
]

CLEAR_ON_CHANGE_PATTERNS = [
    re.compile(r'\bchange[sd]?\b[\s\S]{0,150}\bclear\b', re.IGNORECASE),
    re.compile(r'\bclear(?:ed|s)?\b[\s\S]{0,150}\bchange[sd]?\b', re.IGNORECASE),
    re.compile(r'\bsystem\s+clear[sd]?\b', re.IGNORECASE),
]


def _has_dependency(logic: str) -> re.Match | None:
    for pattern in DEPENDENCY_PATTERNS:
        m = pattern.search(logic)
        if m:
            return m
    return None


def _has_clear_on_change(logic: str) -> bool:
    return any(p.search(logic) for p in CLEAR_ON_CHANGE_PATTERNS)


def _extract_referenced_field(match: re.Match, logic: str, known_fields: set[str]) -> str | None:
    if match.lastindex and match.group(1):
        ref = match.group(1).strip().lower()
        if ref in known_fields:
            return match.group(1).strip()
        for field_name in known_fields:
            if field_name in ref:
                return field_name

    logic_lower = logic.lower()
    for field_name in known_fields:
        if field_name in logic_lower:
            return field_name
    return None


def validate_clear_field_logic(parsed) -> list[ClearFieldLogicResult]:
    """Check that derived fields specify clear-on-change behavior."""
    # Build set of all known field names
    known_fields: set[str] = set()
    for key in ("4.4", "4.5.1", "4.5.2"):
        sec = parsed.sections.get(key)
        if sec:
            known_fields.update(sec.fields_by_name.keys())

    # Map section keys to their field dicts
    section_map = {}
    for key in ("4.4", "4.5.1", "4.5.2"):
        sec = parsed.sections.get(key)
        if sec and sec.has_fields:
            section_map[key] = sec.fields_dict

    results: list[ClearFieldLogicResult] = []

    for section_label, fields in section_map.items():
        for field_name, field_data in fields.items():
            logic = (field_data.get("logic") or "").strip()
            if not logic:
                continue

            match = _has_dependency(logic)
            if match is None:
                continue

            referenced_field = _extract_referenced_field(match, logic, known_fields)

            if not _has_clear_on_change(logic):
                results.append(ClearFieldLogicResult(
                    section=section_label,
                    field_name=field_name,
                    referenced_field=referenced_field or "(detected dependency)",
                    condition_text=match.group(0),
                    logic=logic,
                    status="FAIL",
                    suggestion=(
                        'Please add a clear-on-change clause to the logic '
                        '(e.g., "On change of [source field], clear the field value").'
                    ),
                ))

    return results


# ── Registry integration ────────────────────────────────────────────────────

from openpyxl import Workbook
from .registry import BaseValidator, ValidatorRegistry, ValidationContext
from excel_writer import write_header, apply_severity_fill, auto_width, write_pass_row


@ValidatorRegistry.register
class ClearFieldLogicValidator(BaseValidator):
    name = "Validate Clear Field Logic"
    sheet_name = "Clear Field Logic"
    description = "Checks that fields derived from other fields specify clear-on-change behavior."

    def validate(self, ctx: ValidationContext) -> list[ClearFieldLogicResult]:
        return validate_clear_field_logic(ctx.parsed)

    def write_sheet(self, wb: Workbook, results: list[ClearFieldLogicResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        headers = [
            "Section", "Field Name",
            "Referenced Field", "Condition Found", "Logic", "Status", "Suggestion",
        ]
        write_header(ws, headers)

        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.section)
            ws.cell(row=i, column=2, value=r.field_name)
            ws.cell(row=i, column=3, value=r.referenced_field)
            ws.cell(row=i, column=4, value=r.condition_text)
            ws.cell(row=i, column=5, value=r.logic)
            ws.cell(row=i, column=6, value=r.status)
            apply_severity_fill(ws.cell(row=i, column=6), r.status)
            ws.cell(row=i, column=7, value=r.suggestion)

        if not results:
            write_pass_row(
                ws, 2, len(headers),
                "PASS - All derived fields include a clear-on-change clause.",
            )

        auto_width(ws)
