"""
Check that when a field's logic references another field from a different
panel, the logic also mentions the panel name where the referenced field lives.

Two detection passes:
  1. Quote-based — extracts double-quoted strings from logic.
  2. Name-scan — scans other panels' logic/rules for field name occurrences.
"""

import re
from models import CrossPanelReferenceResult
from section_parser import FieldRow


# ── Helpers ────────────────────────────────────────────────────────────────

def _extract_quoted_field_names(logic: str) -> list[str]:
    matches = re.findall(r'"([^"]+)"', logic)
    result = []
    for m in matches:
        m = m.strip()
        if not m:
            continue
        if re.match(r'^[A-Z0-9_]+$', m) and ' ' not in m:
            continue
        result.append(m)
    return result


def _build_field_to_panel_map(fields: list[FieldRow]) -> dict[str, str]:
    result: dict[str, str] = {}
    for f in fields:
        if f.field_type_upper == "PANEL":
            continue
        result[f.name_lower] = f.panel
    return result


def _get_field_text(f: FieldRow) -> str:
    parts = []
    if f.logic and f.logic.strip():
        parts.append(f.logic)
    if f.rules and f.rules.strip():
        parts.append(f.rules)
    return "\n".join(parts)


def _is_multi_word(name: str) -> bool:
    return len(name.split()) > 1


def _field_name_in_text(field_name: str, text: str) -> bool:
    escaped = re.escape(field_name)
    if _is_multi_word(field_name):
        return bool(re.search(escaped, text, re.IGNORECASE))
    else:
        pattern = r'"[^"]*\b' + escaped + r'\b[^"]*"'
        return bool(re.search(pattern, text, re.IGNORECASE))


def _is_only_part_of_longer_name(
    short_name: str, text: str, all_field_names: list[str]
) -> bool:
    longer_names = [
        n for n in all_field_names
        if len(n) > len(short_name) and short_name.lower() in n.lower()
    ]
    if not longer_names:
        return False
    scrubbed = text
    for ln in sorted(longer_names, key=len, reverse=True):
        scrubbed = re.sub(re.escape(ln), "", scrubbed, flags=re.IGNORECASE)
    escaped = re.escape(short_name)
    return not bool(re.search(escaped, scrubbed, re.IGNORECASE))


# Phrases that indicate an actual field reference rather than descriptive use.
# Checked as regexes — {esc} is replaced with the escaped field name at call time.
_REF_PATTERNS_BEFORE = [
    r'(?:derived|copied?|populated?|fetched?|taken?|picked?)\s+from\s+(?:\w+\s+){{0,3}}{esc}',
    r'(?:same|similar)\s+(?:as|to)\s+(?:\w+\s+){{0,3}}{esc}',
    r'refer(?:s|red|ence|ring)?\s+(?:to\s+)?(?:\w+\s+){{0,3}}{esc}',
    r'value\s+of\s+(?:\w+\s+){{0,3}}{esc}',
    r'(?:copy|copies|copying)\s+(?:to|from)\s+(?:\w+\s+){{0,3}}{esc}',
    r'(?:map(?:ped|s)?|link(?:ed|s)?)\s+(?:to|from|with)\s+(?:\w+\s+){{0,3}}{esc}',
]
_REF_PATTERNS_AFTER = [
    r'{esc}["\']?\s+field\b',
    r'{esc}["\']?\s+value\b',
    r'{esc}["\']?\s+column\b',
    r'{esc}["\']?\s+dropdown\b',
]
_REF_PATTERN_EXACT_QUOTED = r'"\s*{esc}\s*"'


def _has_reference_context(field_name: str, text: str) -> bool:
    """True if *field_name* appears near a referencing phrase in *text*.

    For single-word field names (e.g. "Address") that also happen to be
    common English words, a bare occurrence inside descriptive prose like
    "address will be populated based on GST address" is NOT a genuine
    cross-panel reference.  This helper returns True only when the name
    appears in a clearly referencing context — e.g. ``derived from Address``,
    ``Address field``, or as a standalone quoted string ``"Address"``.
    """
    esc = re.escape(field_name)
    # Exact standalone quoted reference — always counts
    if re.search(_REF_PATTERN_EXACT_QUOTED.format(esc=esc), text, re.IGNORECASE):
        return True
    for pat in _REF_PATTERNS_BEFORE:
        if re.search(pat.format(esc=esc), text, re.IGNORECASE):
            return True
    for pat in _REF_PATTERNS_AFTER:
        if re.search(pat.format(esc=esc), text, re.IGNORECASE):
            return True
    return False


# ── Pass 1: quote-based ──────────────────────────────────────────────────

def _pass_quote_based(
    section_label: str,
    panels: dict[str, list[FieldRow]],
    field_to_panel: dict[str, str],
    all_field_names: set[str],
    all_original_names: list[str],
) -> list[CrossPanelReferenceResult]:
    results: list[CrossPanelReferenceResult] = []

    for panel_name, panel_fields in panels.items():
        panel_field_names = {f.name_lower for f in panel_fields}

        for f in panel_fields:
            logic = f.logic or ""
            if not logic.strip():
                continue

            for ref_name in _extract_quoted_field_names(logic):
                ref_key = ref_name.strip().lower()
                if ref_key not in all_field_names:
                    continue
                if ref_key in panel_field_names:
                    continue

                ref_panel = field_to_panel[ref_key]
                if ref_panel.lower() in logic.lower():
                    continue
                if _is_only_part_of_longer_name(ref_name, logic, all_original_names):
                    continue
                # Single-word names that are common English words need a
                # referencing phrase nearby to count as a real field reference.
                if not _is_multi_word(ref_name) and not _has_reference_context(ref_name, logic):
                    continue

                results.append(CrossPanelReferenceResult(
                    section=section_label,
                    field_name=f.name,
                    panel=panel_name,
                    referenced_field=ref_name,
                    referenced_field_panel=ref_panel,
                    status="FAIL",
                    message=(
                        f'Logic references "{ref_name}" which belongs to '
                        f'panel "{ref_panel}", but the panel name is not '
                        f"mentioned in the logic."
                    ),
                    suggestion=f'Please add "{ref_panel}" panel as reference in the logic.',
                ))

    return results


# ── Pass 2: name-scan ───────────────────────────────────────────────────

def _pass_name_scan(
    section_label: str,
    panels: dict[str, list[FieldRow]],
    field_to_panel: dict[str, str],
    seen_keys: set[tuple],
    all_original_names: list[str],
) -> list[CrossPanelReferenceResult]:
    results: list[CrossPanelReferenceResult] = []

    panel_field_map: dict[str, dict[str, str]] = {}
    for panel_name, panel_fields in panels.items():
        names: dict[str, str] = {}
        for f in panel_fields:
            if f.name:
                names[f.name] = f.name_lower
        panel_field_map[panel_name] = names

    panel_names_list = list(panels.keys())

    for src_panel in panel_names_list:
        src_fields = panel_field_map.get(src_panel, {})
        if not src_fields:
            continue

        for other_panel in panel_names_list:
            if other_panel == src_panel:
                continue

            for f in panels[other_panel]:
                text = _get_field_text(f)
                if not text.strip():
                    continue

                text_lower = text.lower()

                for orig_name, norm_key in src_fields.items():
                    if len(orig_name) <= 2:
                        continue
                    if norm_key == f.name_lower:
                        continue

                    dedup = (section_label, f.name_lower, other_panel.lower(), norm_key)
                    if dedup in seen_keys:
                        continue

                    if not _field_name_in_text(orig_name, text):
                        continue
                    if _is_only_part_of_longer_name(orig_name, text, all_original_names):
                        continue
                    if not _is_multi_word(orig_name) and not _has_reference_context(orig_name, text):
                        continue
                    if src_panel.lower() in text_lower:
                        continue

                    seen_keys.add(dedup)
                    results.append(CrossPanelReferenceResult(
                        section=section_label,
                        field_name=f.name,
                        panel=other_panel,
                        referenced_field=orig_name,
                        referenced_field_panel=src_panel,
                        status="FAIL",
                        message=(
                            f'Logic/rules references "{orig_name}" which belongs to '
                            f'panel "{src_panel}", but the panel name is not '
                            f"mentioned in the logic."
                        ),
                        suggestion=f'Please add "{src_panel}" panel as reference in the logic.',
                    ))

    return results


# ── Main entry point ──────────────────────────────────────────────────────

def validate_cross_panel_references(parsed) -> list[CrossPanelReferenceResult]:
    """Check that cross-panel field references include the panel name."""
    results: list[CrossPanelReferenceResult] = []

    for section_key in ("4.4", "4.5.1", "4.5.2"):
        section = parsed.sections.get(section_key)
        if section is None or not section.has_fields:
            continue

        panels = section.panels
        non_panel_fields = [f for f in section.field_rows if f.field_type_upper != "PANEL"]
        field_to_panel = _build_field_to_panel_map(non_panel_fields)
        all_field_names = set(field_to_panel.keys())

        all_original_names = [f.name for f in non_panel_fields if f.name]

        # Pass 1: quote-based
        pass1 = _pass_quote_based(section_key, panels, field_to_panel, all_field_names, all_original_names)
        results.extend(pass1)

        seen_keys: set[tuple] = set()
        for r in pass1:
            seen_keys.add((r.section, r.field_name.lower(), r.panel.lower(), r.referenced_field.lower()))

        # Pass 2: name-scan
        pass2 = _pass_name_scan(section_key, panels, field_to_panel, seen_keys, all_original_names)
        results.extend(pass2)

    return results


# ── Registry integration ────────────────────────────────────────────────────

from openpyxl import Workbook
from .registry import BaseValidator, ValidatorRegistry, ValidationContext
from excel_writer import write_header, apply_severity_fill, auto_width, write_pass_row


@ValidatorRegistry.register
class CrossPanelReferenceValidator(BaseValidator):
    name = "Validate Cross Panel References"
    sheet_name = "Cross-Panel References"
    description = "Checks that logic referencing fields in other panels includes the target panel name."

    def validate(self, ctx: ValidationContext) -> list[CrossPanelReferenceResult]:
        return validate_cross_panel_references(ctx.parsed)

    def write_sheet(self, wb: Workbook, results: list[CrossPanelReferenceResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        headers = ["Section", "Field Name", "Panel", "Referenced Field", "Referenced Field Panel", "Message", "Status"]
        write_header(ws, headers)

        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.section)
            ws.cell(row=i, column=2, value=r.field_name)
            ws.cell(row=i, column=3, value=r.panel)
            ws.cell(row=i, column=4, value=r.referenced_field)
            ws.cell(row=i, column=5, value=r.referenced_field_panel)
            ws.cell(row=i, column=6, value=r.message)
            ws.cell(row=i, column=7, value=r.status)
            apply_severity_fill(ws.cell(row=i, column=7), r.status)

        if not results:
            write_pass_row(ws, 2, len(headers), "PASS - All cross-panel field references include the panel name.")

        auto_width(ws)
