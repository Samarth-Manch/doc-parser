"""
Fields of certain types must reference a table and column in their logic.
"""

import re
from models import EDVLogicResult

# Strict types: always check for reference table + attribute/column
STRICT_FIELD_TYPES: list[str] = [
    "EXTERNAL_DROP_DOWN_VALUE",
    "DROPDOWN",
    "DROP-DOWN",
]

# Conditional types: only check when EDV is explicitly referenced in the logic
CONDITIONAL_FIELD_TYPES: list[str] = [
    "TEXT",
]


def _check_dropdown_fields(fields: dict, label: str) -> list[EDVLogicResult]:
    results = []
    for name, data in fields.items():
        type_upper = (data.get("type") or "").upper()
        is_strict = type_upper in STRICT_FIELD_TYPES
        is_conditional = type_upper in CONDITIONAL_FIELD_TYPES
        if not is_strict and not is_conditional:
            continue

        logic = data.get("logic") or ""

        # Match explicit keywords (case-insensitive) OR an ALL_CAPS_UNDERSCORE identifier
        has_ref_table = bool(
            re.search(r"Reference|Table|EDV", logic, re.IGNORECASE)
            or re.search(r"\b[A-Z][A-Z0-9]*(?:_[A-Z0-9]+)+\b", logic)
        )
        has_attr_col = bool(re.search(r"attribute|column", logic, re.IGNORECASE))

        # Edge case: logic like "fetched from : TABLE_NAME" is self-contained
        has_fetch_pattern = bool(
            re.search(r"(?:fetched|populated|picked)\s+from\s*[:\-]?\s*\b[A-Z][A-Z0-9]*(?:_[A-Z0-9]+)+\b", logic, re.IGNORECASE)
        )
        if has_fetch_pattern and has_ref_table:
            continue

        if is_conditional:
            has_edv_ref = bool(re.search(r"\bEDV\b", logic, re.IGNORECASE))
            if has_edv_ref and not has_attr_col:
                results.append(EDVLogicResult(
                    section=label,
                    field_name=name,
                    field_type=data.get("type", ""),
                    message="Missing attribute/column in logic",
                    suggestion="Please add the missing reference to the logic.",
                ))
            continue

        if not (has_ref_table and has_attr_col):
            missing = []
            if not has_ref_table:
                missing.append("reference table")
            if not has_attr_col:
                missing.append("attribute/column")
            results.append(EDVLogicResult(
                section=label,
                field_name=name,
                field_type=data.get("type", ""),
                message=f"Missing {' and '.join(missing)} in logic",
                suggestion="Please add the missing reference to the logic.",
            ))

    return results


def validate_external_dropdown_logic(parsed) -> list[EDVLogicResult]:
    """Fields in CHECKED_FIELD_TYPES must reference a table and column."""
    master = parsed.sections.get("4.4")
    if master is None:
        return []

    results = _check_dropdown_fields(master.fields_dict, "4.4")

    skip_sections = {"4.5.1", "4.5.2"}
    for key in ("4.5.1", "4.5.2"):
        sec = parsed.sections.get(key)
        if sec and sec.has_fields and key not in skip_sections:
            results.extend(_check_dropdown_fields(sec.fields_dict, key))

    return results


# ── Registry integration ────────────────────────────────────────────────────

from openpyxl import Workbook
from .registry import BaseValidator, ValidatorRegistry, ValidationContext
from excel_writer import write_header, apply_severity_fill, auto_width, write_pass_row


@ValidatorRegistry.register
class EDVLogicValidator(BaseValidator):
    name = "Validate EDV Logic"
    sheet_name = "EDV Logic"
    description = "Validates that fields of checked types (EDV, DROPDOWN, TEXT, etc.) reference a table and attribute column."

    def validate(self, ctx: ValidationContext) -> list[EDVLogicResult]:
        return validate_external_dropdown_logic(ctx.parsed)

    def write_sheet(self, wb: Workbook, results: list[EDVLogicResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        headers = ["Section", "Field Name", "Field Type", "Message", "Status", "Suggestion"]
        write_header(ws, headers)

        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.section)
            ws.cell(row=i, column=2, value=r.field_name)
            ws.cell(row=i, column=3, value=r.field_type)
            ws.cell(row=i, column=4, value=r.message)
            ws.cell(row=i, column=5, value=r.status)
            apply_severity_fill(ws.cell(row=i, column=5), r.status)
            ws.cell(row=i, column=6, value=r.suggestion)

        if not results:
            write_pass_row(ws, 2, len(headers), "PASS - All EDV/DROPDOWN fields have proper references.")

        auto_width(ws)
