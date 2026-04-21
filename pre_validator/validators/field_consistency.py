"""
Verify every field in each sub-table exists in the master table,
check visibility consistency, and mandatory+invisible logic.
"""

import re

from models import FieldConsistencyResult, VisibilityConsistencyResult, MandatoryInvisibleResult
from section_parser import FIELD_SECTIONS


def validate_field_consistency(parsed) -> list[FieldConsistencyResult]:
    """Verify every field in each sub-table (4.5.1, 4.5.2) exists in the master table (4.4)."""
    master = parsed.sections.get("4.4")
    if master is None or not master.has_fields:
        return []

    master_names = set(master.fields_by_name.keys())
    results = []

    for section_key in ("4.5.1", "4.5.2"):
        section = parsed.sections.get(section_key)
        if section is None or not section.heading_found:
            continue
        if not section.has_fields:
            continue

        sub_names = set(section.fields_by_name.keys())
        missing = sub_names - master_names

        if not missing:
            results.append(FieldConsistencyResult(
                section=section_key, status="PASS", missing_fields="",
                suggestion="No changes needed.",
            ))
        else:
            results.append(FieldConsistencyResult(
                section=section_key, status="FAIL",
                missing_fields=", ".join(sorted(missing)),
                suggestion="Please add the missing fields to the master table (4.4), or remove them from the sub-table.",
            ))

    return results


# ── Visibility helpers ───────────────────────────────────────────────────────

def _logic_indicates_visible(logic: str) -> bool | None:
    logic_lower = logic.strip().lower()
    if not logic_lower:
        return None
    if re.search(r'\binvisible\b', logic_lower) or re.search(r'\bnot\s+visible\b', logic_lower):
        return False
    if re.search(r'\bvisible\b', logic_lower):
        return True
    return None


_CONDITIONAL_PATTERNS = [re.compile(p) for p in [
    r'\bif\b', r'\bwhen\b', r'\bbased\s+on\b',
    r'\bdepending\b', r'\bonly\s+for\b', r'\bcondition\b',
    r'\bat\s+the\s+time\s+of\b',
]]


def _is_conditional_visibility(logic: str) -> bool:
    logic_lower = logic.strip().lower()
    return any(p.search(logic_lower) for p in _CONDITIONAL_PATTERNS)


_DERIVATION_PATTERNS = [re.compile(p) for p in [
    r'\bset\s+to\b', r'\bdefault\b', r'\bderiv(e|ed|es|ing)\b',
    r'\bpopulat(e|ed|es|ing)\b', r'\bauto[-\s]?(fill|populat|generat|assign)\b',
    r'\bassign(ed|s)?\b', r'\breference\b',
    r':=', r'(?<!=)=(?!=)',
]]


def _logic_derives_value(logic: str) -> bool:
    logic_lower = logic.strip().lower()
    return any(p.search(logic_lower) for p in _DERIVATION_PATTERNS)


_REMOVES_MANDATORY_PATTERNS = [re.compile(p) for p in [
    r'\bnon[\s-]?mandatory\b',
    r'\bnot[\s-]?mandatory\b',
    r'\bmandatory\s*[:\s]\s*no\b',
    r'\bremove\s+(the\s+)?mandatory\b',
    r'\bmade\s+invisible\s+and\s+non[\s-]?mandatory\b',
    r'\binvisible\s+and\s+non[\s-]?mandatory\b',
    r'\bnon[\s-]?mandatory\s+and\s+(non[\s-]?)?invisible\b',
]]


def _logic_removes_mandatory_when_invisible(logic: str) -> bool:
    logic_lower = logic.strip().lower()
    return any(p.search(logic_lower) for p in _REMOVES_MANDATORY_PATTERNS)


# ── Visibility consistency ───────────────────────────────────────────────────

def validate_visibility_consistency(parsed) -> list[VisibilityConsistencyResult]:
    """If a field in 4.4 has logic indicating 'visible', it should exist in sub-tables."""
    master = parsed.sections.get("4.4")
    if master is None or not master.has_fields:
        return []

    # Collect available sub-tables
    available: dict[str, set[str]] = {}
    for key in ("4.5.1", "4.5.2"):
        sec = parsed.sections.get(key)
        if sec and sec.has_fields:
            available[key] = set(sec.fields_by_name.keys())

    if not available:
        return []

    results = []
    for f in master.field_rows:
        if f.field_type_upper == "PANEL":
            continue
        visibility = _logic_indicates_visible(f.logic)
        if visibility is not True:
            continue

        present_in_any = any(f.name_lower in names for names in available.values())
        missing_sections = [
            key for key, names in available.items()
            if f.name_lower not in names
        ]

        if not present_in_any:
            severity = "WARNING" if _is_conditional_visibility(f.logic) else "FAIL"
            results.append(VisibilityConsistencyResult(
                field_name=f.name,
                logic=f.logic,
                status=severity,
                missing_in=", ".join(missing_sections),
                suggestion="Please add the missing field to any of the sub-tables, or remove it from the master table, or mark it as invisible.",
            ))

    return results


# ── Mandatory invisible ─────────────────────────────────────────────────────

def validate_mandatory_invisible(parsed) -> list[MandatoryInvisibleResult]:
    """If a field in 4.4 is invisible AND mandatory, the logic should derive a value."""
    master = parsed.sections.get("4.4")
    if master is None or not master.has_fields:
        return []

    results = []
    for f in master.field_rows:
        if f.field_type_upper == "PANEL":
            continue
        visibility = _logic_indicates_visible(f.logic)
        if visibility is not False:
            continue
        if not f.is_mandatory:
            continue

        if _logic_removes_mandatory_when_invisible(f.logic):
            results.append(MandatoryInvisibleResult(
                field_name=f.name, logic=f.logic, is_mandatory=True,
                status="PASS",
                reason="Logic removes mandatory status when field is invisible",
                suggestion="No changes needed.",
            ))
        elif _logic_derives_value(f.logic):
            results.append(MandatoryInvisibleResult(
                field_name=f.name, logic=f.logic, is_mandatory=True,
                status="PASS",
                reason="Logic derives a value for the invisible mandatory field",
                suggestion="No changes needed.",
            ))
        else:
            results.append(MandatoryInvisibleResult(
                field_name=f.name, logic=f.logic, is_mandatory=True,
                status="FAIL",
                reason="Field is invisible and mandatory but logic does not derive a value",
                suggestion="Please make the field non-mandatory, or add logic that derives a value for it.",
            ))

    return results


# ── Registry integration ────────────────────────────────────────────────────

from openpyxl import Workbook
from .registry import BaseValidator, ValidatorRegistry, ValidationContext
from excel_writer import write_header, apply_severity_fill, auto_width, write_pass_row


@ValidatorRegistry.register
class FieldConsistencyValidator(BaseValidator):
    name = "Validate Field Consistency"
    sheet_name = "Field Consistency"
    description = "Checks that every field in sub-tables (4.5.1/4.5.2) exists in the master table (4.4)."

    def validate(self, ctx: ValidationContext) -> list[FieldConsistencyResult]:
        return validate_field_consistency(ctx.parsed)

    def write_sheet(self, wb: Workbook, results: list[FieldConsistencyResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        headers = ["Section", "Missing Fields", "Status", "Suggestion"]
        write_header(ws, headers)

        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.section)
            ws.cell(row=i, column=2, value=r.missing_fields)
            ws.cell(row=i, column=3, value=r.status)
            apply_severity_fill(ws.cell(row=i, column=3), r.status)
            ws.cell(row=i, column=4, value=r.suggestion)

        auto_width(ws)


@ValidatorRegistry.register
class VisibilityConsistencyValidator(BaseValidator):
    name = "Validate Visibility Consistency"
    sheet_name = "Visibility Consistency"
    description = "Ensures field visibility settings are consistent between master and sub-tables."

    def validate(self, ctx: ValidationContext) -> list[VisibilityConsistencyResult]:
        return validate_visibility_consistency(ctx.parsed)

    def write_sheet(self, wb: Workbook, results: list[VisibilityConsistencyResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        headers = ["Field Name", "Logic", "Missing In Sections", "Status", "Suggestion"]
        write_header(ws, headers)

        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.field_name)
            ws.cell(row=i, column=2, value=r.logic)
            ws.cell(row=i, column=3, value=r.missing_in)
            ws.cell(row=i, column=4, value=r.status)
            apply_severity_fill(ws.cell(row=i, column=4), r.status)
            ws.cell(row=i, column=5, value=r.suggestion)

        if not results:
            write_pass_row(ws, 2, len(headers), "PASS - No visible fields with missing sub-table entries.")

        auto_width(ws)


@ValidatorRegistry.register
class MandatoryInvisibleValidator(BaseValidator):
    name = "Validate Invisible Mandatory Consistency"
    sheet_name = "Mandatory Invisible"
    description = "Flags fields that are both mandatory and invisible, which is a logical conflict."

    def validate(self, ctx: ValidationContext) -> list[MandatoryInvisibleResult]:
        return validate_mandatory_invisible(ctx.parsed)

    def write_sheet(self, wb: Workbook, results: list[MandatoryInvisibleResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        headers = ["Field Name", "Logic", "Mandatory", "Reason", "Status", "Suggestion"]
        write_header(ws, headers)

        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.field_name)
            ws.cell(row=i, column=2, value=r.logic)
            ws.cell(row=i, column=3, value="TRUE" if r.is_mandatory else "FALSE")
            ws.cell(row=i, column=4, value=r.reason)
            ws.cell(row=i, column=5, value=r.status)
            apply_severity_fill(ws.cell(row=i, column=5), r.status)
            ws.cell(row=i, column=6, value=r.suggestion)

        if not results:
            write_pass_row(ws, 2, len(headers), "PASS - No mandatory invisible fields without value derivation.")

        auto_width(ws)
