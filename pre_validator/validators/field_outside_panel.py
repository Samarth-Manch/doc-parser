"""
Every field must belong to a panel. Fields that appear before the first
PANEL row (or in a table with no PANEL rows at all) are flagged as errors.
"""

from models import FieldOutsidePanelResult


def validate_fields_outside_panel(parsed) -> list[FieldOutsidePanelResult]:
    """Check that no field exists outside a panel in any section."""
    results: list[FieldOutsidePanelResult] = []

    for section_label in ("4.4", "4.5.1", "4.5.2"):
        section = parsed.sections.get(section_label)
        if section is None or not section.heading_found or not section.field_rows:
            results.append(FieldOutsidePanelResult(
                section=section_label, field_name="",
                message="Section not found or has no fields.",
                status="N/A", suggestion="No changes needed.",
            ))
            continue

        found_error = False
        for f in section.field_rows:
            if f.field_type_upper == "PANEL":
                continue
            if not f.panel:
                found_error = True
                results.append(FieldOutsidePanelResult(
                    section=section_label, field_name=f.name,
                    message=f"Field '{f.name}' is not inside any panel.",
                    status="FAIL",
                    suggestion="Please add a PANEL field type before this field, or move this field inside an existing panel.",
                ))

        if not found_error:
            results.append(FieldOutsidePanelResult(
                section=section_label, field_name="",
                message="All fields are inside a panel.",
                status="PASS", suggestion="No changes needed.",
            ))

    return results


# ── Registry integration ────────────────────────────────────────────────────

from openpyxl import Workbook
from .registry import BaseValidator, ValidatorRegistry, ValidationContext
from excel_writer import write_header, apply_severity_fill, auto_width, write_pass_row


@ValidatorRegistry.register
class FieldOutsidePanelValidator(BaseValidator):
    name = "Validate Fields Outside Panel"
    sheet_name = "Fields Outside Panel"
    description = "Ensures every field belongs to a panel; flags fields outside any panel."

    def validate(self, ctx: ValidationContext) -> list[FieldOutsidePanelResult]:
        return validate_fields_outside_panel(ctx.parsed)

    def write_sheet(self, wb: Workbook, results: list[FieldOutsidePanelResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        headers = ["Section", "Field Name", "Message", "Status", "Suggestion"]
        write_header(ws, headers)

        has_issues = any(r.status == "FAIL" for r in results)

        if not has_issues:
            write_pass_row(ws, 2, len(headers), "PASS - All fields are inside a panel.")
        else:
            row = 2
            for r in results:
                ws.cell(row=row, column=1, value=r.section)
                ws.cell(row=row, column=2, value=r.field_name if r.field_name else "\u2014")
                ws.cell(row=row, column=3, value=r.message)
                ws.cell(row=row, column=4, value=r.status)
                apply_severity_fill(ws.cell(row=row, column=4), r.status)
                ws.cell(row=row, column=5, value=r.suggestion)
                row += 1

        auto_width(ws)
