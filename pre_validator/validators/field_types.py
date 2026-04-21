"""
Verify every value in the 'Field Type' column is from the allowed list.
Uses fuzzy matching (Levenshtein + token overlap) to suggest corrections.
"""

import re

from rapidfuzz.distance import Levenshtein
from models import FieldTypeResult
from section_parser import FIELD_SECTIONS


VALID_FIELD_TYPES = {
    "TEXT", "IMAGE", "RADIAL_BUTTON", "CHECK_BOX", "GRP_HDR", "GRP_END",
    "OPTION", "DATE", "TEXTAREA", "PANEL", "COMMENT", "FORMULA", "VARIABLE",
    "PAN", "VOTER_ID", "PAN_NUMBER", "BUTTON_HDR", "BUTTON_END", "BUTTON_ICON",
    "ACTION", "OTP", "VIDEO_NATIVE", "AUDIO", "AUDIO_OTP", "AUDIO_FORM_FILL",
    "STATIC_CHECK_BOX", "STATIC_LOCATION", "DYNAMIC_LOCATION",
    "EXTERNAL_DROP_DOWN_VALUE", "EXTERNAL_DROP_DOWN_MULTISELECT",
    "MULTISELECT_EXTERNAL_DROPDOWN", "TIME", "QR_SCANNER",
    "MASKED_FIELD", "FOUR_DIGITS", "IFRAME", "VIDEO", "FILE", "MULTIPLE_FILE",
    "VIDEO_KYC", "VIDEO_KYC_RECORD", "VIDEO_AGENT_KYC", "CONTENT_VALUE_INFO",
    "COMPASS_DIRECTION", "DYNAMIC_BUTTON", "PREVIEW", "HTML_PREVIEW",
    "IMAGE_DISPLAY", "ROW_HDR", "ROW_END", "LABEL", "EXECUTE_BUTTON",
    "MAKE_PAYMENT", "CHECK_PAYMENT", "MAKE_ESTAMP", "NOTES", "NOTE_HISTORY",
    "TABLE_VIEW", "IMAGE_VIEW", "PASSWORD", "EXTERNAL_DROP_DOWN_RADIOBUTTON",
    "AUDIO_RECORD", "ARRAY_HDR", "ARRAY_END", "PDF", "ADVANCE_MAP",
    "CARD_HDR", "CARD_END", "BASE_DOC_VIEW", "OTP_BOX", "DROPDOWN",
    "MOBILE", "EMAIL",
}

VALID_FIELD_TYPES_LIST = sorted(VALID_FIELD_TYPES)

UNSUPPORTED_TYPE_REPLACEMENTS = {"NUMBER": "TEXT", "NUMERIC": "TEXT"}

FUZZY_THRESHOLD = 0.90

_MULTI_TYPE_SEP = re.compile(r"[,/|;\n]+")


def _lev_similarity(a: str, b: str) -> float:
    return Levenshtein.normalized_similarity(a, b)


def _tokenize(s: str) -> list[str]:
    return [t for t in re.split(r'[\s_\-]+', s.upper()) if t]


def _token_score(input_tokens: list[str], candidate_tokens: list[str]) -> float:
    if not input_tokens:
        return 0.0
    matched = sum(
        1 for tok in input_tokens
        if any(ct == tok or tok in ct or ct in tok for ct in candidate_tokens)
    )
    return matched / len(input_tokens)


def _combined_score(input_str: str, candidate: str) -> float:
    lev = _lev_similarity(input_str, candidate)
    t_score = _token_score(_tokenize(input_str), _tokenize(candidate))
    return max(lev, t_score)


def _best_fuzzy_match(input_str: str) -> dict:
    upper = input_str.upper()
    if upper in VALID_FIELD_TYPES:
        return {"match": upper, "score": 1.0, "exact": True}

    best_match = None
    best_score = 0.0
    for valid_type in VALID_FIELD_TYPES_LIST:
        score = _combined_score(upper, valid_type)
        if score > best_score:
            best_score = score
            best_match = valid_type

    return {"match": best_match, "score": best_score, "exact": False}


def _looks_like_multiple_types(raw: str) -> bool:
    parts = [p.strip() for p in _MULTI_TYPE_SEP.split(raw) if p.strip()]
    if len(parts) < 2:
        return False
    recognised = sum(1 for p in parts if p.upper() in VALID_FIELD_TYPES)
    return recognised >= 2


def check_field_types(parsed) -> list[FieldTypeResult]:
    """Validate every 'Field Type' cell against the allowed list."""
    results = []

    for section_key in FIELD_SECTIONS:
        section = parsed.sections.get(section_key)
        if section is None or not section.heading_found:
            continue

        for f in section.field_rows:
            raw_type = f.field_type
            field_name = f.name[:50]

            # Blank/empty field type
            if not raw_type:
                results.append(FieldTypeResult(
                    section=section_key, field_name=field_name,
                    invalid_type="(blank)", status="FAIL",
                    suggestion="Field type is missing, please add a valid field type.",
                ))
                continue

            # Explicitly unsupported types
            replacement = UNSUPPORTED_TYPE_REPLACEMENTS.get(raw_type.upper())
            if replacement is not None:
                results.append(FieldTypeResult(
                    section=section_key, field_name=field_name,
                    invalid_type=raw_type, status="FAIL",
                    suggestion=(
                        f'"{raw_type.upper()}" is not supported, please replace with "{replacement}" '
                        "(a regex is applied on this field in the backend)."
                    ),
                ))
                continue

            # Multiple types in a single cell
            if _looks_like_multiple_types(raw_type):
                results.append(FieldTypeResult(
                    section=section_key, field_name=field_name,
                    invalid_type=raw_type, status="WARNING",
                    suggestion="Given field type is not supported, please use a single valid field type.",
                ))
                continue

            # Fuzzy matching
            fuzzy = _best_fuzzy_match(raw_type)
            if fuzzy["exact"]:
                continue

            status = "WARNING" if fuzzy["score"] >= FUZZY_THRESHOLD else "FAIL"
            results.append(FieldTypeResult(
                section=section_key, field_name=field_name,
                invalid_type=raw_type, status=status,
                suggestion=f'Given field type is not supported, please replace with "{fuzzy["match"]}".',
            ))

    return results


# ── Registry integration ────────────────────────────────────────────────────

from openpyxl import Workbook
from .registry import BaseValidator, ValidatorRegistry, ValidationContext
from excel_writer import write_header, apply_severity_fill, auto_width, write_pass_row


@ValidatorRegistry.register
class FieldTypeValidator(BaseValidator):
    name = "Validate Field Types"
    sheet_name = "Field Types"
    description = "Checks that every value in the Field Type column matches an allowed BUD type (fuzzy, 90% threshold)."

    def validate(self, ctx: ValidationContext) -> list[FieldTypeResult]:
        return check_field_types(ctx.parsed)

    def write_sheet(self, wb: Workbook, results: list[FieldTypeResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        headers = ["Section", "Field Name", "Invalid Type", "Status", "Suggestion"]
        write_header(ws, headers)

        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.section)
            ws.cell(row=i, column=2, value=r.field_name)
            ws.cell(row=i, column=3, value=r.invalid_type)
            ws.cell(row=i, column=4, value=r.status)
            apply_severity_fill(ws.cell(row=i, column=4), r.status)
            ws.cell(row=i, column=5, value=r.suggestion)

        if not results:
            write_pass_row(ws, 2, len(headers), "PASS - All field types are valid (single, recognized types).")

        auto_width(ws)
