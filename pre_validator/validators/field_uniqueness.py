"""
Every field name must be unique within each panel of each table.
A panel is defined by a row whose type column is PANEL.

Reads field_rows from the parsed document (which preserves all raw rows
including consecutive duplicates that doc_parser would merge).
"""

from collections import Counter, defaultdict
from models import FieldUniquenessResult, FieldDuplicateRow, PanelUniquenessRow
from section_parser import FIELD_SECTIONS


# Field types that denote ARRAY boundary markers.
_ARRAY_BOUNDARY_TYPES = {"ARRAY_HDR", "ARRAY_END", "ARRAY HEADER", "ARRAY END"}


def _find_duplicates(rows: list[tuple[str, str]]) -> list[str]:
    """Return field names that appear more than once.

    Excludes duplicates where every occurrence is an ARRAY boundary marker.
    """
    grouped: dict[str, list[str]] = defaultdict(list)
    for name, ftype in rows:
        grouped[name].append(ftype)

    duplicates = []
    for name, types in grouped.items():
        if len(types) <= 1:
            continue
        if all(t.strip().upper() in _ARRAY_BOUNDARY_TYPES for t in types if t):
            continue
        duplicates.append(name)
    return sorted(duplicates)


def validate_field_uniqueness(parsed) -> FieldUniquenessResult:
    """Check that every field name is unique within each panel."""
    field_duplicates: list[FieldDuplicateRow] = []
    panel_uniqueness: list[PanelUniquenessRow] = []

    for section_key in ("4.4", "4.5.1", "4.5.2"):
        section = parsed.sections.get(section_key)
        if section is None or not section.heading_found or not section.field_rows:
            field_duplicates.append(FieldDuplicateRow(
                section=section_key, panel="N/A", field_name="", status="N/A",
                suggestion="Section has no fields, no changes needed.",
            ))
            panel_uniqueness.append(PanelUniquenessRow(
                section=section_key, duplicate_panels="", status="N/A",
                suggestion="Section has no fields, no changes needed.",
            ))
            continue

        # --- Field name uniqueness (Table 1) ---
        # Group non-PANEL rows by panel
        panels_rows: dict[str, list[tuple[str, str]]] = defaultdict(list)
        for f in section.field_rows:
            if f.field_type_upper == "PANEL":
                continue
            panels_rows[f.panel].append((f.name_lower, f.field_type))

        for panel_name, panel_rows in panels_rows.items():
            dupes = _find_duplicates(panel_rows)
            if dupes:
                for dup in dupes:
                    field_duplicates.append(FieldDuplicateRow(
                        section=section_key, panel=panel_name,
                        field_name=dup, status="FAIL",
                        suggestion="Please delete the duplicate field or change the field name.",
                    ))
            else:
                field_duplicates.append(FieldDuplicateRow(
                    section=section_key, panel=panel_name,
                    field_name="", status="PASS",
                    suggestion="No changes needed.",
                ))

        # --- Panel name uniqueness (Table 2) ---
        panel_names = [
            f.name_lower for f in section.field_rows
            if f.field_type_upper == "PANEL"
        ]
        dup_panels = {name for name, count in Counter(panel_names).items() if count > 1}

        if dup_panels:
            panel_uniqueness.append(PanelUniquenessRow(
                section=section_key,
                duplicate_panels=", ".join(sorted(dup_panels)),
                status="FAIL",
                suggestion="Please delete the duplicate panel or change the panel name.",
            ))
        else:
            panel_uniqueness.append(PanelUniquenessRow(
                section=section_key, duplicate_panels="", status="PASS",
                suggestion="No changes needed.",
            ))

    return FieldUniquenessResult(
        field_duplicates=field_duplicates,
        panel_uniqueness=panel_uniqueness,
    )


# ── Registry integration ────────────────────────────────────────────────────

from openpyxl import Workbook
from .registry import BaseValidator, ValidatorRegistry, ValidationContext
from excel_writer import write_header, apply_severity_fill, auto_width, write_pass_row


@ValidatorRegistry.register
class FieldUniquenessValidator(BaseValidator):
    name = "Validate Field Uniqueness"
    sheet_name = "Field Uniqueness"
    description = "Ensures field names are unique within each panel and that panel names are not duplicated."

    def validate(self, ctx: ValidationContext) -> list[FieldUniquenessResult]:
        return [validate_field_uniqueness(ctx.parsed)]

    def write_sheet(self, wb: Workbook, results: list[FieldUniquenessResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        result = results[0]

        # ── Table 1: Field Name Uniqueness ──
        headers1 = ["Section", "Panel", "Duplicate Field", "Status", "Suggestion"]
        write_header(ws, headers1)
        row = 2

        for r in result.field_duplicates:
            ws.cell(row=row, column=1, value=r.section)
            ws.cell(row=row, column=2, value=r.panel)
            ws.cell(row=row, column=3, value=r.field_name if r.field_name else ("\u2014" if r.status == "PASS" else ""))
            ws.cell(row=row, column=4, value=r.status)
            apply_severity_fill(ws.cell(row=row, column=4), r.status)
            ws.cell(row=row, column=5, value=r.suggestion)
            row += 1

        # ── Blank separator row ──
        row += 1

        # ── Table 2: Panel Name Uniqueness ──
        has_panel_duplicates = any(r.status == "FAIL" for r in result.panel_uniqueness)

        if has_panel_duplicates:
            headers2 = ["Section", "Duplicate Panel Names", "Status", "Suggestion"]
            from excel_writer import HEADER_FONT, HEADER_FILL
            from openpyxl.styles import Alignment
            for col, header in enumerate(headers2, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for r in result.panel_uniqueness:
                if r.status == "FAIL":
                    ws.cell(row=row, column=1, value=r.section)
                    ws.cell(row=row, column=2, value=r.duplicate_panels)
                    ws.cell(row=row, column=3, value=r.status)
                    apply_severity_fill(ws.cell(row=row, column=3), r.status)
                    ws.cell(row=row, column=4, value=r.suggestion)
                    row += 1
        else:
            write_pass_row(ws, row, 4, "PASS - No duplicate panels.")

        auto_width(ws)
