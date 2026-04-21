"""
Checks that fields marked as non-editable in the Logic column have
their Field Type set to "TEXT" or "DATE".
"""

from models import NonEditableCheckResult
from section_parser import FIELD_SECTIONS


NON_EDITABLE_PHRASES = [
    "always non editable",
    "always non-editable",
    "non-editable",
    "non editable",
]

ALLOWED_NON_EDITABLE_TYPES = {"TEXT", "DATE"}


def validate_non_editable_fields(parsed) -> list[NonEditableCheckResult]:
    """Check that non-editable fields have allowed types."""
    results: list[NonEditableCheckResult] = []

    for section_key in FIELD_SECTIONS:
        section = parsed.sections.get(section_key)
        if section is None or not section.heading_found:
            continue

        for f in section.field_rows:
            if f.field_type_upper == "PANEL":
                continue
            if not f.logic or not f.field_type:
                continue

            logic_lower = f.logic.lower()
            matched_phrase = None
            for phrase in NON_EDITABLE_PHRASES:
                if phrase in logic_lower:
                    matched_phrase = phrase
                    break

            if matched_phrase and f.field_type_upper not in ALLOWED_NON_EDITABLE_TYPES:
                results.append(NonEditableCheckResult(
                    section=section_key,
                    field_name=f.name[:50],
                    field_type=f.field_type,
                    status="FAIL",
                    suggestion=f"{f.field_type} cannot be non-editable, make sure that field should not be non-editable.",
                ))

    return results


# ── Registry integration ────────────────────────────────────────────────────

from openpyxl import Workbook
from .registry import BaseValidator, ValidatorRegistry, ValidationContext
from excel_writer import write_header, apply_severity_fill, auto_width, write_pass_row


@ValidatorRegistry.register
class NonEditableValidator(BaseValidator):
    name = "Validate Non-Editable Field Types"
    sheet_name = "Non-Editable Check"
    description = "Verifies that fields marked as non-editable have their Field Type set to TEXT."

    def validate(self, ctx: ValidationContext) -> list[NonEditableCheckResult]:
        return validate_non_editable_fields(ctx.parsed)

    def write_sheet(self, wb: Workbook, results: list[NonEditableCheckResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        headers = ["Section", "Field Name", "Field Type", "Status", "Suggestion"]
        write_header(ws, headers)

        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.section)
            ws.cell(row=i, column=2, value=r.field_name)
            ws.cell(row=i, column=3, value=r.field_type)
            ws.cell(row=i, column=4, value=r.status)
            apply_severity_fill(ws.cell(row=i, column=4), r.status)
            ws.cell(row=i, column=5, value=r.suggestion)

        if not results:
            write_pass_row(ws, 2, len(headers), "PASS - All non-editable fields have type TEXT.")

        auto_width(ws)
