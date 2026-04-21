"""
Validates Section 4.7: Record List View.

Checks that the list of fields displayed on dashboards contains more
than just the default sample fields.
"""

from models import RecordListViewResult


SAMPLE_FIELDS = [
    "Transaction ID",
    "Company Name \u2013 Title",
    "Process Name",
    "Create By",
    "Created Date",
    "Status",
]

_DASHES = {"\u2013", "\u2014", "-", "\u2013", "\u2014"}


def _match_sample(field_text: str) -> str | None:
    for sample in SAMPLE_FIELDS:
        if field_text == sample:
            return sample
        if field_text.startswith(sample):
            rest = field_text[len(sample):].lstrip()
            if not rest or rest[0] in _DASHES or rest[0] == "(":
                return sample
    return None


def validate_record_list_view(parsed) -> list[RecordListViewResult]:
    """Validate Section 4.7 Record List View fields against the sample list."""
    results: list[RecordListViewResult] = []

    section = parsed.sections.get("4.7")
    if section is None or not section.heading_found:
        results.append(RecordListViewResult(
            message="Section 4.7 Record List View not found in the document.",
            status="WARNING",
            suggestion="Please add section 4.7 Record List View to the document.",
        ))
        return results

    fields = parsed.section_47_fields
    if not fields:
        results.append(RecordListViewResult(
            message="No fields found under Section 4.7.",
            status="WARNING",
            suggestion="Please add the record list view fields under section 4.7.",
        ))
        return results

    # "Type" is not allowed
    if any(f.strip().lower() == "type" for f in fields):
        results.append(RecordListViewResult(
            message='A field named "Type" is not allowed in Section 4.7 Record List View.',
            status="FAIL",
            suggestion='Please remove or rename the field named "Type" in Section 4.7.',
        ))

    matched_samples: set[str] = set()
    extra_fields: list[str] = []

    for f in fields:
        sample_match = _match_sample(f)
        if sample_match:
            matched_samples.add(sample_match)
        else:
            extra_fields.append(f)

    missing_sample = sorted(set(SAMPLE_FIELDS) - matched_samples)

    if not extra_fields and not missing_sample:
        results.append(RecordListViewResult(
            message=(
                "The Record List View contains only the default sample fields. "
                "Please review and add any additional fields specific to your process."
            ),
            status="WARNING",
            suggestion="Please add fields specific to your process along with the sample fields.",
        ))
    elif not extra_fields and missing_sample:
        results.append(RecordListViewResult(
            message=(
                "The Record List View contains only sample fields and is missing: "
                + ", ".join(missing_sample) + ". "
                "Please review and add any additional fields specific to your process."
            ),
            status="WARNING",
            suggestion="Please add the missing sample fields and any fields specific to your process.",
        ))
    elif extra_fields and missing_sample:
        results.append(RecordListViewResult(
            message=(
                "The Record List View has additional fields but is missing these "
                "sample fields: " + ", ".join(missing_sample) + ". "
                "Please verify whether these are intentionally excluded."
            ),
            status="WARNING",
            suggestion="Please add the missing sample fields, or confirm they are intentionally excluded.",
        ))
    else:
        results.append(RecordListViewResult(
            message="The Record List View contains all sample fields and additional custom fields.",
            status="PASS",
            suggestion="No changes needed.",
        ))

    return results


# ── Registry integration ────────────────────────────────────────────────────

from openpyxl import Workbook
from .registry import BaseValidator, ValidatorRegistry, ValidationContext
from excel_writer import write_header, apply_severity_fill, auto_width, write_pass_row


@ValidatorRegistry.register
class RecordListViewValidator(BaseValidator):
    name = "Validate Record List View"
    sheet_name = "Record List View"
    description = "Checks that Section 4.7 Record List View contains more than just sample fields."

    def validate(self, ctx: ValidationContext) -> list[RecordListViewResult]:
        return validate_record_list_view(ctx.parsed)

    def write_sheet(self, wb: Workbook, results: list[RecordListViewResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        headers = ["Message", "Status", "Suggestion"]
        write_header(ws, headers)

        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.message)
            ws.cell(row=i, column=2, value=r.status)
            apply_severity_fill(ws.cell(row=i, column=2), r.status)
            ws.cell(row=i, column=3, value=r.suggestion)

        if not results:
            write_pass_row(ws, 2, len(headers), "PASS - Record List View fields look good.")

        auto_width(ws)
