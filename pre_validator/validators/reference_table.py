"""
Validates that reference tables cited in field logic exist in section 4.6.

Uses the pre-parsed section_46_table_names from BUDDocument (populated by
section_parser) and regex-extracts table references from field logic.
"""

import json
import logging
import os
import re
import subprocess
from models import ReferenceTableResult
from section_parser import normalise_table_name, normalise_table_name_variants

# ── Toggle LLM fallback ─────────────────────────────────────────────────────
USE_LLM = True

# ── Patterns for extracting table references FROM field logic ────────────────

_PAT_NUMBERED = re.compile(
    r"(?:refer(?:ence)?\s+)?table\s+(\d+\.\d+)(?:\s*\(([A-Za-z_][A-Za-z0-9_]*)\))?",
    re.IGNORECASE,
)
_PAT_NAMED_AFTER_SEP = re.compile(
    r"(?:reference\s+)?table\s*(?:name)?\s*[-:]\s*([A-Za-z_][A-Za-z0-9_]*)",
    re.IGNORECASE,
)
_PAT_NAMED_AFTER_SPACE = re.compile(
    r'reference\s+table\s+["\u201c\u201d\u201e]?([A-Za-z_][A-Za-z0-9_]*(?:_[A-Za-z0-9_]+)+)["\u201c\u201d\u201f]?',
    re.IGNORECASE,
)
_PAT_NAMED_QUOTED = re.compile(
    r'(?:refer(?:ence)?\s+)?table\s+["\u201c\u201e]([A-Za-z_][A-Za-z0-9_ ]*?)["\u201d\u201f]',
    re.IGNORECASE,
)
_PAT_NAME_BEFORE = re.compile(
    r"(\b[A-Za-z_][A-Za-z0-9_]*)\s+reference\s+table",
    re.IGNORECASE,
)
_PAT_EDV_QUOTED = re.compile(
    r'\bEDV[,;]?\s+(?:(?:refer(?:ence)?\s+)?table\s+)?["\u201c\u201e]([A-Z][A-Z0-9_ ]+?)["\u201d\u201f]',
)
_PAT_EDV_UNQUOTED = re.compile(
    r'\bEDV[,;]?\s+(?:(?:refer(?:ence)?\s+)?table\s+)?([A-Z][A-Z0-9_]+)(?=\s|$|[,;.])',
)
_PAT_QUOTED_IDENTIFIER = re.compile(
    r'["\u201c\u201e]([A-Z][A-Z0-9]*(?:_[A-Z0-9]+)+)["\u201d\u201f]',
)
_PAT_COLUMN_OF_QUOTED = re.compile(
    r'column\s+\d+\s+of\s+(?:the\s+)?(?:(?:refer(?:ence)?\s+)?table\s+)?'
    r'["\u201c\u201e]([A-Za-z_][A-Za-z0-9_ ]*?)["\u201d\u201f]',
    re.IGNORECASE,
)
_PAT_COLUMN_OF_UNQUOTED = re.compile(
    r'column\s+\d+\s+of\s+(?:the\s+)?(?:(?:refer(?:ence)?\s+)?table\s+)?'
    r'([A-Z][A-Z0-9]*(?:_[A-Z0-9_]+)+)(?=\s|$|[,;.])',
    re.IGNORECASE,
)

_NON_TABLE_KEYWORDS = frozenset({"edv", "staging", "data", "table", "of", "the", "reference", "refer"})


def _is_table_identifier(name: str, raw_text: str) -> bool:
    if name.lower() in _NON_TABLE_KEYWORDS:
        return False
    if "_" in name:
        return True
    if raw_text.isupper() and len(raw_text) >= 2:
        return True
    return False


def _extract_logic_refs(logic: str) -> dict[str, str]:
    """Return {normalised_name: original_name} for table refs in a logic string."""
    refs: dict[str, str] = {}

    def add_ref(raw: str) -> None:
        key = normalise_table_name(raw)
        if key not in refs:
            refs[key] = raw.strip()

    for m in _PAT_NUMBERED.finditer(logic):
        add_ref(m.group(1))
        if m.group(2):
            add_ref(m.group(2))
    for m in _PAT_NAMED_AFTER_SEP.finditer(logic):
        if m.group(1).lower() not in _NON_TABLE_KEYWORDS:
            add_ref(m.group(1))
    for m in _PAT_NAMED_AFTER_SPACE.finditer(logic):
        add_ref(m.group(1))
    for m in _PAT_NAMED_QUOTED.finditer(logic):
        raw = m.group(1).strip()
        if raw.lower() not in _NON_TABLE_KEYWORDS:
            add_ref(raw)
    for m in _PAT_NAME_BEFORE.finditer(logic):
        if _is_table_identifier(m.group(1), m.group(1)):
            add_ref(m.group(1))
    for m in _PAT_EDV_QUOTED.finditer(logic):
        add_ref(m.group(1).strip())
    for m in _PAT_EDV_UNQUOTED.finditer(logic):
        raw = m.group(1).strip()
        if normalise_table_name(raw) not in refs:
            add_ref(raw)
    for m in _PAT_COLUMN_OF_QUOTED.finditer(logic):
        raw = m.group(1).strip()
        if raw.lower() not in _NON_TABLE_KEYWORDS:
            add_ref(raw)
    for m in _PAT_COLUMN_OF_UNQUOTED.finditer(logic):
        raw = m.group(1).strip()
        if raw.lower() not in _NON_TABLE_KEYWORDS and normalise_table_name(raw) not in refs:
            add_ref(raw)
    for m in _PAT_QUOTED_IDENTIFIER.finditer(logic):
        raw = m.group(1).strip()
        if raw.lower() not in _NON_TABLE_KEYWORDS and normalise_table_name(raw) not in refs:
            add_ref(raw)

    return refs


# ── Field checker ────────────────────────────────────────────────────────────

def _check_fields(
    fields: dict, section_label: str, available_normalised: set[str],
) -> list[ReferenceTableResult]:
    results: list[ReferenceTableResult] = []
    already_reported: set[tuple[str, str]] = set()

    for field_name, data in fields.items():
        logic = data.get("logic") or ""
        for norm_ref, orig_ref in _extract_logic_refs(logic).items():
            if norm_ref in available_normalised:
                continue
            key = (field_name, norm_ref)
            if key in already_reported:
                continue
            already_reported.add(key)
            display = f"Table {orig_ref}" if re.match(r"\d+\.\d+$", norm_ref) else orig_ref
            results.append(ReferenceTableResult(
                section=section_label, field_name=field_name,
                referenced_table=display, status="WARNING",
                message=f"'{display}' referenced in logic but not found in section 4.6 Reference Tables",
                suggestion=f'Please make sure you have referenced the correct EDV table, or add "{display}" in section 4.6.',
            ))

    return results


# ── LLM fallback ────────────────────────────────────────────────────────────

_EDV_DROPDOWN_TYPES = frozenset({
    "EXTERNAL_DROP_DOWN_VALUE", "DROPDOWN", "DROP-DOWN",
    "EXTERNAL_DROPDOWN_VALUE", "EDV",
})
_CONDITIONAL_TYPES = frozenset({"TEXT"})
LLM_BATCH_SIZE = 15


def _is_claude_cli_available() -> bool:
    try:
        result = subprocess.run(
            ["claude", "--version"], capture_output=True, text=True, timeout=10,
        )
        return result.returncode == 0
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def _collect_llm_candidates(
    fields: dict, section_label: str, available_normalised: set[str],
    deterministic_warned: set[tuple[str, str]],
) -> list[dict]:
    candidates = []
    for field_name, data in fields.items():
        logic = data.get("logic") or ""
        if not logic.strip():
            continue
        field_type = (data.get("type") or "").upper()
        is_edv_dropdown = field_type in _EDV_DROPDOWN_TYPES
        is_conditional = field_type in _CONDITIONAL_TYPES
        if not is_edv_dropdown and not is_conditional:
            continue
        refs = _extract_logic_refs(logic)
        unmatched_refs = {n: o for n, o in refs.items() if n not in available_normalised}
        if unmatched_refs:
            candidates.append({
                "section": section_label, "field_name": field_name,
                "field_type": data.get("type", ""), "logic": logic[:500],
                "category": "A", "unmatched_refs": list(unmatched_refs.values()),
            })
        elif not refs:
            candidates.append({
                "section": section_label, "field_name": field_name,
                "field_type": data.get("type", ""), "logic": logic[:500],
                "category": "B", "unmatched_refs": [],
            })
    return candidates


def _llm_resolve_batch(candidates: list[dict], available_table_names: list[str]) -> dict[tuple[str, str], dict]:
    if not candidates:
        return {}

    field_descriptions = []
    for idx, c in enumerate(candidates):
        desc = (
            f"Field {idx + 1}:\n"
            f"  Name: {c['field_name']}\n  Type: {c['field_type']}\n"
            f"  Section: {c['section']}\n  Logic: \"{c['logic']}\"\n"
        )
        if c["unmatched_refs"]:
            desc += f"  Regex-extracted refs (not found in 4.6): {c['unmatched_refs']}\n"
        field_descriptions.append(desc)

    table_list = "\n".join(f"  - {name}" for name in sorted(available_table_names))
    prompt = (
        "You are a BUD (Business Understanding Document) validator. "
        "I will give you fields from a BUD document whose logic text may or may not "
        "reference an external data source (EDV table, staging table, reference table, "
        "lookup table, master data, etc.). Our deterministic regex could not detect "
        "a table reference for these fields.\n\n"
        "Your task has TWO steps for each field:\n"
        "1. DETECT: Read the logic carefully and determine whether it references an "
        "EDV / reference table / staging table that should exist in section 4.6.\n\n"
        "IMPORTANT — The following are NOT EDV/reference table references:\n"
        "  - External service / API validations (GSTIN, PAN, MSME, CIN, bank verification)\n"
        "  - Field-to-field derivations from validation responses\n"
        "  - Generic descriptors ('master data', 'system generated')\n\n"
        "2. MATCH: If found, match to the 4.6 list below.\n\n"
        f"Available tables in section 4.6:\n{table_list}\n\n"
        "Respond ONLY with a JSON array:\n"
        '  [{"field": N, "references_table": bool, "detected_name": str|null, '
        '"matched_table": str|null, "reason": str}]\n\n'
        "Fields:\n" + "\n".join(field_descriptions)
    )

    try:
        result = subprocess.run(
            ["claude", "-p", prompt, "--output-format", "text", "--model", "claude-haiku-4-5-20251001"],
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode != 0:
            return {}
        text = result.stdout.strip()
        if "```" in text:
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        results = json.loads(text)
    except Exception:
        return {}

    resolved = {}
    for item in results:
        field_idx = item.get("field", 0) - 1
        if 0 <= field_idx < len(candidates):
            c = candidates[field_idx]
            resolved[(c["section"], c["field_name"])] = {
                "references_table": item.get("references_table", False),
                "detected_name": item.get("detected_name"),
                "matched_table": item.get("matched_table"),
                "reason": item.get("reason", ""),
            }
    return resolved


def _run_llm_fallback(
    all_fields: dict[str, tuple[dict, str]],
    available_normalised: set[str],
    deterministic_results: list[ReferenceTableResult],
) -> list[ReferenceTableResult]:
    if not _is_claude_cli_available():
        logging.warning("Claude CLI not available. Skipping LLM reference table fallback.")
        return deterministic_results

    det_warned = {(r.section, r.field_name) for r in deterministic_results if r.status == "WARNING"}

    candidates = []
    for section_label, (fields, _) in all_fields.items():
        candidates.extend(_collect_llm_candidates(fields, section_label, available_normalised, det_warned))

    if not candidates:
        return deterministic_results

    available_list = sorted(available_normalised)
    logging.info("Sending %d field(s) to Claude CLI for reference table matching...", len(candidates))

    all_resolved: dict[tuple[str, str], dict] = {}
    for batch_start in range(0, len(candidates), LLM_BATCH_SIZE):
        batch = candidates[batch_start:batch_start + LLM_BATCH_SIZE]
        all_resolved.update(_llm_resolve_batch(batch, available_list))

    logging.info("LLM reference table matching complete.")

    final_results = []
    for r in deterministic_results:
        key = (r.section, r.field_name)
        if r.status == "WARNING" and key in all_resolved:
            if all_resolved[key]["matched_table"]:
                continue
        final_results.append(r)

    for c in candidates:
        if c["category"] != "B":
            continue
        key = (c["section"], c["field_name"])
        llm_result = all_resolved.get(key, {})
        if llm_result.get("matched_table"):
            continue
        references_table = llm_result.get("references_table", False)
        detected_name = llm_result.get("detected_name")
        reason = llm_result.get("reason", "")
        if references_table:
            display = detected_name or "(name unclear)"
            final_results.append(ReferenceTableResult(
                section=c["section"], field_name=c["field_name"],
                referenced_table=display, status="WARNING",
                message=f"Logic references '{display}' but it was not found in section 4.6 (detected by LLM)",
                suggestion=f'Please add "{display}" in section 4.6.' + (f" LLM note: {reason}" if reason else ""),
            ))
        elif c["field_type"].upper() in _EDV_DROPDOWN_TYPES:
            final_results.append(ReferenceTableResult(
                section=c["section"], field_name=c["field_name"],
                referenced_table="(not detected)", status="WARNING",
                message=f"Field type is {c['field_type']} but no reference table could be identified in logic",
                suggestion="Please specify the EDV/reference table name clearly in the logic text."
                + (f" LLM note: {reason}" if reason else ""),
            ))

    return final_results


# ── Public API ───────────────────────────────────────────────────────────────

def validate_reference_tables(parsed, use_llm: bool = True) -> list[ReferenceTableResult]:
    """Check that every reference table cited in field logic is listed in section 4.6."""
    available = parsed.section_46_table_names

    if not available:
        return [ReferenceTableResult(
            section="4.6", field_name="N/A", referenced_table="N/A",
            status="WARNING",
            message="Section 4.6 Reference Tables not found or contains no table entries",
            suggestion="Please add section 4.6 with the reference tables used in the document.",
        )]

    # Gather field dicts for each section
    master = parsed.sections.get("4.4")
    master_fields = master.fields_dict if master and master.has_fields else {}

    results = _check_fields(master_fields, "4.4", available)
    for key in ("4.5.1", "4.5.2"):
        sec = parsed.sections.get(key)
        if sec and sec.has_fields:
            results.extend(_check_fields(sec.fields_dict, key, available))

    if not use_llm:
        return results

    all_fields_map = {"4.4": (master_fields, "4.4")}
    for key in ("4.5.1", "4.5.2"):
        sec = parsed.sections.get(key)
        if sec and sec.has_fields:
            all_fields_map[key] = (sec.fields_dict, key)

    results = _run_llm_fallback(all_fields_map, available, results)
    return results


# ── Registry integration ────────────────────────────────────────────────────

from openpyxl import Workbook
from .registry import BaseValidator, ValidatorRegistry, ValidationContext
from excel_writer import write_header, apply_severity_fill, auto_width, write_pass_row


@ValidatorRegistry.register
class ReferenceTableValidator(BaseValidator):
    name = "Validate Reference Tables"
    sheet_name = "Reference Tables"
    description = "Validates that reference tables cited in field logic exist as entries in section 4.6."

    def validate(self, ctx: ValidationContext) -> list[ReferenceTableResult]:
        return validate_reference_tables(ctx.parsed, use_llm=USE_LLM)

    def write_sheet(self, wb: Workbook, results: list[ReferenceTableResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        headers = ["Section", "Field Name", "Referenced Table", "Message", "Status", "Suggestion"]
        write_header(ws, headers)

        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.section)
            ws.cell(row=i, column=2, value=r.field_name)
            ws.cell(row=i, column=3, value=r.referenced_table)
            ws.cell(row=i, column=4, value=r.message)
            ws.cell(row=i, column=5, value=r.status)
            apply_severity_fill(ws.cell(row=i, column=5), r.status)
            ws.cell(row=i, column=6, value=r.suggestion)

        if not results:
            write_pass_row(ws, 2, len(headers), "PASS - All referenced tables exist in section 4.6.")

        auto_width(ws)
