"""
Compare rules (Logic column) for each field in the master table (4.4)
against its matching field in 4.5.1 and/or 4.5.2.

Detection is tiered and deterministic-first:
  Tier 1 — Equality  : exact match after whitespace normalisation     -> WARNING
  Tier 2 — Fuzzy     : rapidfuzz similarity >= FUZZY_THRESHOLD        -> WARNING
  Tier 3 — LLM       : Claude Code CLI for remaining unmatched pairs   -> WARNING
"""

import json
import logging
import re
import subprocess
from dataclasses import dataclass
from rapidfuzz import fuzz
from models import RuleDuplicateResult


# ── Configuration ─────────────────────────────────────────────────────────

FUZZY_THRESHOLD      = 95
MIN_RULE_LENGTH      = 30
LLM_BATCH_SIZE       = 15


@dataclass
class RuleEntry:
    field_name: str
    logic:      str
    section:    str
    row_index:  int


# ── Rule extraction ───────────────────────────────────────────────────────

def _normalize_rule(s: str) -> str:
    return " ".join(s.lower().split())


def _normalize_field_name(name: str) -> str:
    return " ".join(name.lower().split()).strip("/ ")


def extract_rules(parsed) -> dict[str, list[RuleEntry]]:
    """Extract non-trivial rules from field rows in all field sections."""
    sections_map: dict[str, list[RuleEntry]] = {}

    for section_key in ("4.4", "4.5.1", "4.5.2"):
        section = parsed.sections.get(section_key)
        if section is None or not section.heading_found:
            continue

        rules = []
        for f in section.field_rows:
            if f.field_type_upper == "PANEL":
                continue
            logic_text = f.logic.strip()
            if logic_text and logic_text.lower() not in ("", "disable", "disabled"):
                rules.append(RuleEntry(
                    field_name=f.name,
                    logic=logic_text,
                    section=section_key,
                    row_index=f.row_index,
                ))

        sections_map[section_key] = rules

    return sections_map


# ── Tier 1: Equality ──────────────────────────────────────────────────────

def check_equality(rule_a: str, rule_b: str) -> bool:
    return _normalize_rule(rule_a) == _normalize_rule(rule_b)


# ── Tier 2: Fuzzy ─────────────────────────────────────────────────────────

def check_fuzzy(rule_a: str, rule_b: str) -> tuple[bool, float]:
    score = fuzz.ratio(rule_a.lower(), rule_b.lower())
    return score >= FUZZY_THRESHOLD, score


# ── Tier 3: LLM (Claude Code CLI) ────────────────────────────────────────

def _is_claude_cli_available() -> bool:
    try:
        result = subprocess.run(
            ["claude", "--version"],
            capture_output=True, text=True, timeout=10,
        )
        return result.returncode == 0
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def check_llm_batch(
    pairs: list[tuple[RuleEntry, RuleEntry]],
) -> list[tuple[RuleEntry, RuleEntry, str]]:
    if not pairs:
        return []

    pair_descriptions = []
    for idx, (ra, rb) in enumerate(pairs):
        pair_descriptions.append(
            f"Pair {idx + 1}:\n"
            f"  Rule A (field: '{ra.field_name}', sec {ra.section}): \"{ra.logic[:300]}\"\n"
            f"  Rule B (field: '{rb.field_name}', sec {rb.section}): \"{rb.logic[:300]}\""
        )

    prompt = (
        "You are a BUD (Business Understanding Document) validator. I will give you pairs of rules "
        "from different fields/sections in a document. For each pair, determine if they are "
        "semantically duplicate — meaning they express the same business logic even if the "
        "wording differs.\n\n"
        "Respond ONLY with a JSON array. Each element should be an object with:\n"
        '  {"pair": <pair_number>, "duplicate": true/false, "reason": "<brief explanation>"}\n\n'
        "Rules:\n" + "\n\n".join(pair_descriptions)
    )

    try:
        result = subprocess.run(
            ["claude", "-p", prompt, "--output-format", "text"],
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode != 0:
            logging.warning("Claude CLI returned non-zero exit code: %s", result.stderr)
            return []

        text = result.stdout.strip()
        if "```" in text:
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        results = json.loads(text)
    except subprocess.TimeoutExpired:
        logging.warning("Claude CLI call timed out")
        return []
    except json.JSONDecodeError as e:
        logging.warning("Failed to parse Claude CLI response as JSON: %s", e)
        return []
    except Exception as e:
        logging.warning("Claude CLI call failed: %s", e)
        return []

    duplicates = []
    for item in results:
        if item.get("duplicate"):
            pair_idx = item["pair"] - 1
            if 0 <= pair_idx < len(pairs):
                ra, rb = pairs[pair_idx]
                duplicates.append((ra, rb, item.get("reason", "")))

    return duplicates


# ── Validator ─────────────────────────────────────────────────────────────

def check_rule_uniqueness(parsed, use_llm: bool = True) -> list[RuleDuplicateResult]:
    """Check rule uniqueness across sections. Tiers: EXACT -> FUZZY -> LLM."""
    sections_map = extract_rules(parsed)
    if not sections_map:
        return []

    master_rules = sections_map.get("4.4", [])
    if not master_rules:
        return []

    sub_sections = ["4.5.1", "4.5.2"]

    if use_llm and not _is_claude_cli_available():
        logging.warning("Claude CLI not available on PATH. Skipping LLM tier.")
        use_llm = False

    results: list[RuleDuplicateResult] = []
    already_flagged: set[tuple[str, int, str, int]] = set()
    llm_candidates:  list[tuple[RuleEntry, RuleEntry]] = []

    def _flag_key(ra: RuleEntry, rb: RuleEntry):
        a = (ra.section, ra.row_index)
        b = (rb.section, rb.row_index)
        return (*min(a, b), *max(a, b))

    def _compare(rule_a: RuleEntry, rule_b: RuleEntry) -> None:
        key = _flag_key(rule_a, rule_b)
        if key in already_flagged:
            return

        loc = f"Section {rule_a.section} vs Section {rule_b.section}"

        if check_equality(rule_a.logic, rule_b.logic):
            already_flagged.add(key)
            results.append(RuleDuplicateResult(
                status="WARNING", location=loc, detection_tier="EXACT",
                field_a=rule_a.field_name, section_a=rule_a.section,
                rule_a=rule_a.logic[:150],
                field_b=rule_b.field_name, section_b=rule_b.section,
                rule_b=rule_b.logic[:150],
                reason="Rules are exactly the same after normalizing spaces and casing.",
                suggestion=f"Please remove the duplicate rule from section {rule_b.section} as it is the same as in section {rule_a.section}.",
            ))
            return

        is_fuzzy, fuzzy_score = check_fuzzy(rule_a.logic, rule_b.logic)
        if is_fuzzy:
            already_flagged.add(key)
            results.append(RuleDuplicateResult(
                status="WARNING", location=loc, detection_tier="FUZZY",
                field_a=rule_a.field_name, section_a=rule_a.section,
                rule_a=rule_a.logic[:120],
                field_b=rule_b.field_name, section_b=rule_b.section,
                rule_b=rule_b.logic[:120],
                reason=f"Rules are {fuzzy_score:.0f}% similar based on text comparison.",
                suggestion=f"Please review and remove the duplicate rule from section {rule_b.section} if it is the same as in section {rule_a.section}.",
            ))
            return

        if use_llm and len(rule_a.logic) >= MIN_RULE_LENGTH and len(rule_b.logic) >= MIN_RULE_LENGTH:
            llm_candidates.append((rule_a, rule_b))

    for sub_sec in sub_sections:
        sub_rules = sections_map.get(sub_sec, [])
        if not sub_rules:
            continue
        index_sub: dict[str, list[RuleEntry]] = {}
        for rule in sub_rules:
            index_sub.setdefault(_normalize_field_name(rule.field_name), []).append(rule)
        for master_rule in master_rules:
            for sub_rule in index_sub.get(_normalize_field_name(master_rule.field_name), []):
                _compare(master_rule, sub_rule)

    if use_llm and llm_candidates:
        logging.info("Sending %d pair(s) to Claude Code CLI for matching...", len(llm_candidates))
        for batch_start in range(0, len(llm_candidates), LLM_BATCH_SIZE):
            batch = llm_candidates[batch_start:batch_start + LLM_BATCH_SIZE]
            for rule_a, rule_b, reason in check_llm_batch(batch):
                key = _flag_key(rule_a, rule_b)
                if key in already_flagged:
                    continue
                already_flagged.add(key)
                loc = f"Section {rule_a.section} vs Section {rule_b.section}"
                results.append(RuleDuplicateResult(
                    status="WARNING", location=loc, detection_tier="LLM",
                    field_a=rule_a.field_name, section_a=rule_a.section,
                    rule_a=rule_a.logic[:120],
                    field_b=rule_b.field_name, section_b=rule_b.section,
                    rule_b=rule_b.logic[:120],
                    reason=reason,
                    suggestion=f"Please review and remove the duplicate rule from section {rule_b.section} if it is the same as in section {rule_a.section}.",
                ))
        logging.info("LLM matching complete.")

    return results


# ── Registry integration ────────────────────────────────────────────────────

from openpyxl import Workbook
from .registry import BaseValidator, ValidatorRegistry, ValidationContext
from excel_writer import write_header, apply_severity_fill, auto_width, write_pass_row


@ValidatorRegistry.register
class RuleUniquenessValidator(BaseValidator):
    name = "Validate Rule Uniqueness"
    sheet_name = "Rule Uniqueness"
    description = "Detects duplicate rules across sections using exact, fuzzy, and LLM-based matching."

    def validate(self, ctx: ValidationContext) -> list[RuleDuplicateResult]:
        return check_rule_uniqueness(ctx.parsed, use_llm=True)

    def write_sheet(self, wb: Workbook, results: list[RuleDuplicateResult]) -> None:
        ws = wb.create_sheet(self.sheet_name)
        headers = [
            "Location", "Detection Tier",
            "Field A", "Section A", "Rule A",
            "Field B", "Section B", "Rule B",
            "Reason", "Status", "Suggestion",
        ]
        write_header(ws, headers)

        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.location)
            ws.cell(row=i, column=2, value=r.detection_tier)
            ws.cell(row=i, column=3, value=r.field_a)
            ws.cell(row=i, column=4, value=r.section_a)
            ws.cell(row=i, column=5, value=r.rule_a)
            ws.cell(row=i, column=6, value=r.field_b)
            ws.cell(row=i, column=7, value=r.section_b)
            ws.cell(row=i, column=8, value=r.rule_b)
            ws.cell(row=i, column=9, value=r.reason)
            ws.cell(row=i, column=10, value=r.status)
            apply_severity_fill(ws.cell(row=i, column=10), r.status)
            ws.cell(row=i, column=11, value=r.suggestion)

        if not results:
            write_pass_row(ws, 2, len(headers), "PASS - No duplicate rules found.")

        auto_width(ws)
