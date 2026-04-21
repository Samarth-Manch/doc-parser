"""
Verify that sections 4.4, 4.5.1, and 4.5.2 each contain a
properly structured field-level table.
"""

from models import TableCheckResult
from section_parser import FIELD_SECTIONS

# Sections that MUST be present (4.4 = error, others = warning)
_REQUIRED = [
    ("4.4",   "4.4 Field-Level Information"),
    ("4.5.1", "4.5.1 Initiator Behaviour"),
    ("4.5.2", "4.5.2 Vendor Behaviour"),
]


def check_table_existence(parsed) -> list[TableCheckResult]:
    """Verify required sections each contain a proper field-level table."""
    results = []

    for section_num, section_heading in _REQUIRED:
        section = parsed.sections.get(section_num)
        missing_status = "FAIL" if section_num == "4.4" else "WARNING"

        if section is None or not section.heading_found:
            results.append(TableCheckResult(
                status=missing_status,
                section=section_num,
                message=f"Section '{section_num}' not found in document.",
                details="Expected a heading containing this section number.",
                suggestion=f"Please add section {section_num} to the document.",
            ))
            continue

        all_tables = section.raw_tables
        field_tables = [t for t in all_tables if t.is_field_table]

        if not all_tables:
            if section.is_not_applicable:
                results.append(TableCheckResult(
                    status="INFO",
                    section=section_num,
                    message=f"Section '{section_num}' is marked 'Not Applicable' — no table expected.",
                    details=f"Heading: '{section.heading_text}'",
                    suggestion="No changes needed.",
                ))
            else:
                results.append(TableCheckResult(
                    status=missing_status,
                    section=section_num,
                    message=f"Section '{section_num}' exists but contains no tables.",
                    details=f"Heading: '{section.heading_text}'",
                    suggestion='Please add a table to this section, or mark it as "Not Applicable" if it does not apply.',
                ))

        elif not field_tables:
            results.append(TableCheckResult(
                status="WARNING",
                section=section_num,
                message=(
                    f"Section '{section_num}' has {len(all_tables)} table(s) "
                    f"but none match the expected field-table structure."
                ),
                details="Expected columns: 'Field Name', 'Field Type', 'Mandatory', 'Logic'",
                suggestion="Please rewrite the table using the standard column layout: Field Name, Field Type, Mandatory, Logic.",
            ))

        else:
            for ft in field_tables:
                row_count = len(ft.rows)
                if row_count < 1:
                    results.append(TableCheckResult(
                        status="WARNING",
                        section=section_num,
                        message=f"Section '{section_num}' has an empty field table (header row only).",
                        details="Table has no data rows.",
                        suggestion='Please add fields to the table, or mark the section as "Not Applicable" if it does not apply.',
                    ))
                else:
                    results.append(TableCheckResult(
                        status="PASS",
                        section=section_num,
                        message=f"Field table found with {row_count} row(s).",
                        details="",
                        suggestion="No changes needed.",
                    ))

    return results


# ── Registry integration ────────────────────────────────────────────────────

from openpyxl import Workbook
from .registry import BaseValidator, ValidatorRegistry, ValidationContext
from excel_writer import write_header, apply_severity_fill, auto_width, write_pass_row


@ValidatorRegistry.register
class TableFormattingValidator(BaseValidator):
    name = "Validate Table Consistency"
    sheet_name = "Table Formatting"
    description = "Verifies that sections 4.4, 4.5.1, and 4.5.2 each contain a properly structured field-level table."

    def validate(self, ctx: ValidationContext) -> list[TableCheckResult]:
        return check_table_existence(ctx.parsed)

    def write_sheet(self, wb: Workbook, results: list[TableCheckResult]) -> None:
        ws = wb.active
        ws.title = self.sheet_name
        headers = ["Section", "Message", "Details", "Status", "Suggestion"]
        write_header(ws, headers)

        for i, r in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=r.section)
            ws.cell(row=i, column=2, value=r.message)
            ws.cell(row=i, column=3, value=r.details)
            ws.cell(row=i, column=4, value=r.status)
            apply_severity_fill(ws.cell(row=i, column=4), r.status)
            ws.cell(row=i, column=5, value=r.suggestion)

        if not results:
            write_pass_row(ws, 2, len(headers), "PASS - All sections have valid tables.")

        auto_width(ws)
